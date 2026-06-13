from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from books.models import Book, Category


class Command(BaseCommand):
    help = "Import books from an Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default=str(
                Path("data") / "books (1).xlsx"
            ),
            help="Path to an .xlsx file."
        )

        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse the workbook and roll back database changes."
        )

        parser.add_argument(
            "--strict",
            action="store_true",
            help="Fail immediately when a row contains invalid data."
        )

    def handle(self, *args, **options):
        workbook_path = Path(
            options["path"]
        )

        if not workbook_path.exists():
            raise CommandError(
                f"Workbook not found: {workbook_path}"
            )

        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=True
        )

        worksheet = workbook.active
        rows = worksheet.iter_rows(
            values_only=True
        )

        try:
            headers = next(
                rows
            )
        except StopIteration as exc:
            raise CommandError(
                "Workbook is empty."
            ) from exc

        columns = [
            self.normalize_header(
                header
            )
            for header in headers
        ]

        required = {
            "title",
            "authors",
            "publisher",
            "published_date",
            "category",
            "distribution_expense",
        }

        missing = sorted(
            required - set(
                columns
            )
        )

        if missing:
            raise CommandError(
                "Missing required columns: "
                + ", ".join(
                    missing
                )
            )

        created = 0
        updated = 0
        skipped = 0

        with transaction.atomic():
            for index, row in enumerate(
                rows,
                start=2
            ):
                record = dict(
                    zip(
                        columns,
                        row
                    )
                )

                if not self.clean_text(
                    record.get(
                        "title"
                    )
                ):
                    skipped += 1
                    continue

                category_name = self.clean_text(
                    record.get(
                        "category"
                    )
                )

                if not category_name:
                    skipped += 1
                    self.warn_or_raise(
                        options,
                        f"Row {index}: category is required; skipped."
                    )
                    continue

                try:
                    published_date = self.parse_date(
                        record.get(
                            "published_date"
                        ),
                        index
                    )

                    distribution_expense = self.parse_decimal(
                        record.get(
                            "distribution_expense"
                        ),
                        index
                    )
                except CommandError as exc:
                    skipped += 1
                    self.warn_or_raise(
                        options,
                        f"{exc}; skipped."
                    )
                    continue

                category, _ = Category.objects.get_or_create(
                    name=category_name
                )

                defaults = {
                    "title": self.clean_text(
                        record.get(
                            "title"
                        )
                    ),
                    "subtitle": self.clean_text(
                        record.get(
                            "subtitle"
                        )
                    ),
                    "authors": self.clean_text(
                        record.get(
                            "authors"
                        )
                    ),
                    "publisher": self.clean_text(
                        record.get(
                            "publisher"
                        )
                    ),
                    "published_date": published_date,
                    "category": category,
                    "distribution_expense": distribution_expense,
                }

                isbn = self.clean_text(
                    record.get(
                        "isbn"
                    )
                    or record.get(
                        "id"
                    )
                )

                if isbn:
                    defaults["isbn"] = isbn

                _, was_created = Book.objects.update_or_create(
                    title=defaults["title"],
                    authors=defaults["authors"],
                    publisher=defaults["publisher"],
                    defaults=defaults
                )

                if was_created:
                    created += 1
                else:
                    updated += 1

            if options["dry_run"]:
                transaction.set_rollback(
                    True
                )

        mode = "Dry run parsed" if options["dry_run"] else "Imported"

        self.stdout.write(
            self.style.SUCCESS(
                f"{mode}: {created} created, {updated} updated, {skipped} skipped."
            )
        )

    def normalize_header(self, value):
        return str(
            value or ""
        ).strip().lower().replace(
            " ",
            "_"
        )

    def warn_or_raise(self, options, message):
        if options["strict"]:
            raise CommandError(
                message
            )

        self.stderr.write(
            self.style.WARNING(
                message
            )
        )

    def clean_text(self, value):
        if value is None:
            return ""

        return str(
            value
        ).strip()

    def parse_date(self, value, row_number):
        if isinstance(
            value,
            datetime
        ):
            return value.date()

        if isinstance(
            value,
            date
        ):
            return value

        text = self.clean_text(
            value
        )

        normalized_text = text.translate(
            str.maketrans(
                {
                    "i": "1",
                    "I": "1",
                    "l": "1",
                    "O": "0",
                    "o": "0",
                }
            )
        )

        candidates = []

        for candidate_text in (
            normalized_text,
            re.sub(
                r"[^\d/-]",
                "",
                text
            ),
        ):
            date_match = re.search(
                r"\d{1,4}[-/]\d{1,2}[-/]\d{2,4}",
                candidate_text
            )

            if date_match:
                candidates.append(
                    date_match.group(
                        0
                    )
                )

        for candidate in candidates:
            for date_format in (
                "%m/%d/%Y",
                "%Y-%m-%d",
                "%d/%m/%Y",
            ):
                try:
                    return datetime.strptime(
                        candidate,
                        date_format
                    ).date()
                except ValueError:
                    pass

        raise CommandError(
            f"Row {row_number}: invalid published_date '{value}'."
        )

    def parse_decimal(self, value, row_number):
        try:
            return Decimal(
                str(
                    value
                ).strip()
            )
        except (InvalidOperation, AttributeError) as exc:
            raise CommandError(
                f"Row {row_number}: invalid distribution_expense '{value}'."
            ) from exc
