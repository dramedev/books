import base64
import csv
import hashlib
import hmac as _hmac
import json
import logging
import math
import random
import re
import secrets
import stripe
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, update_session_auth_hash
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.exceptions import PermissionDenied
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import Case, Count, DecimalField, ExpressionWrapper, F, IntegerField, Max, OuterRef, Q, Subquery, Sum, Value, When
from django.db.models.functions import Coalesce, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.http import url_has_allowed_host_and_scheme
from django.urls import reverse
from django.utils import timezone, translation
from django.utils.translation import gettext, gettext_lazy
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from . import ai_chat, iyzico_client
from .isbn_lookup import IsbnLookupError, lookup_isbn
from .analytics import PURCHASE_COST_EXPRESSION, REVENUE_EXPRESSION
from .reorder_logic import (
    REORDER_VELOCITY_WINDOW_DAYS, REORDER_COVER_DAYS,
    daily_sales_velocity as _daily_sales_velocity,
    suggested_reorder_quantity as _suggested_reorder_quantity,
)
from .forms import (
    AcceptInviteForm,
    AuthorForm,
    BookForm,
    CategoryForm,
    CustomerForm,
    IntegrationForm,
    InviteUserForm,
    InvoiceForm,
    InvoiceItemForm,
    IyzicoCustomerForm,
    LocationForm,
    PrintRunForm,
    EmailUpdateForm,
    ProfileForm,
    RedeemAccessCodeForm,
    ReorderForm,
    ReturnForm,
    RoyaltyPaymentForm,
    RoyaltyRateForm,
    SaleForm,
    SignupForm,
    StockAdjustmentForm,
    StockTransferForm,
    SupplierForm,
    VerifyEmailForm,
)
from .models import (
    CURRENCY_CHOICES,
    AccessCode, Account, AccountInvitation, AccountMembership, Author, Book, Category, Customer, CustomerLoginToken,
    Integration, Invoice, InvoiceItem,
    Location, PrintRun, ProcessedShopifyOrder, Profile,
    Reorder, Return, RoyaltyPayment, RoyaltyRate,
    Sale, StockAdjustment, StockLevel, Subscription, Supplier,
)
from .permissions import ensure_roles, sync_user_groups_for_role


logger = logging.getLogger(__name__)


def _csv_safe(value):
    """Neutralize CSV/formula injection: a leading =, +, -, @, tab or CR lets
    a poisoned cell run as a formula in Excel/LibreOffice when the export is
    later opened. Prefixing with a single quote forces it to be read as text."""
    text = str(value)
    if text and text[0] in "=+-@\t\r":
        return "'" + text
    return text


def _book_export_headers():
    return [
        gettext("ISBN"),
        gettext("Title"),
        gettext("Subtitle"),
        gettext("Authors"),
        gettext("Publisher"),
        gettext("Published Date"),
        gettext("Category"),
        gettext("Distribution Expense"),
    ]


LEARNING_QUOTES = [
    # Learning & growth
    gettext_lazy('"The beautiful thing about learning is that no one can take it away from you." — B.B. King'),
    gettext_lazy('"Live as if you were to die tomorrow. Learn as if you were to live forever." — Mahatma Gandhi'),
    gettext_lazy('"An investment in knowledge always pays the best interest." — Benjamin Franklin'),
    gettext_lazy('"The capacity to learn is a gift; the ability to learn is a skill; the willingness to learn is a choice." — Brian Herbert'),
    gettext_lazy('"Develop a passion for learning. If you do, you will never cease to grow." — Anthony J. D\'Angelo'),
    gettext_lazy('"Each small task of everyday life is part of the total harmony of the universe." — Saint Therese'),
    gettext_lazy('"Growth is painful. Change is painful. But nothing is as painful as staying stuck somewhere you don\'t belong." — N.R. Narayana Murthy'),
    gettext_lazy('"The expert in anything was once a beginner." — Helen Hayes'),
    gettext_lazy('"Success is the sum of small efforts repeated day in and day out." — Robert Collier'),
    gettext_lazy('"You don\'t have to be great to start, but you have to start to be great." — Zig Ziglar'),

    # Books & reading
    gettext_lazy('"A room without books is like a body without a soul." — Marcus Tullius Cicero'),
    gettext_lazy('"Books are a uniquely portable magic." — Stephen King'),
    gettext_lazy('"Today a reader, tomorrow a leader." — Margaret Fuller'),
    gettext_lazy('"Reading is to the mind what exercise is to the body." — Joseph Addison'),
    gettext_lazy('"Once you learn to read, you will be forever free." — Frederick Douglass'),
    gettext_lazy('"I have always imagined that Paradise will be a kind of library." — Jorge Luis Borges'),
    gettext_lazy('"So many books, so little time." — Frank Zappa'),
    gettext_lazy('"There is no friend as loyal as a book." — Ernest Hemingway'),
    gettext_lazy('"A reader lives a thousand lives before he dies. The man who never reads lives only one." — George R.R. Martin'),
    gettext_lazy('"Books are mirrors: you only see in them what you already have inside you." — Carlos Ruiz Zafón'),

    # Book distribution mission
    gettext_lazy('"Every book has a destination, and every reader has a journey." — Book Distribution Philosophy'),
    gettext_lazy('"We do not simply move books; we move knowledge, ideas, and imagination." — Book Distribution Philosophy'),
    gettext_lazy('"A warehouse full of books is a warehouse full of possibilities." — Book Distribution Philosophy'),
    gettext_lazy('"Every delivered book is a new story beginning somewhere." — Book Distribution Philosophy'),
    gettext_lazy('"Behind every order is a reader waiting for discovery." — Book Distribution Philosophy'),
    gettext_lazy('"Distribution turns printed pages into shared experiences." — Book Distribution Philosophy'),
    gettext_lazy('"Every package carries imagination, knowledge, and opportunity." — Book Distribution Philosophy'),
    gettext_lazy('"A distributor is the bridge between authors and readers." — Book Distribution Philosophy'),
    gettext_lazy('"The journey of knowledge begins with accessibility." — Book Distribution Philosophy'),
    gettext_lazy('"Books travel so minds can explore." — Book Distribution Philosophy'),

    # Business & teamwork
    gettext_lazy('"The goal as a company is to have customer service that is not just the best but legendary." — Sam Walton'),
    gettext_lazy('"Quality is the best business plan." — John Lasseter'),
    gettext_lazy('"Great things in business are never done by one person. They are done by a team of people." — Steve Jobs'),
    gettext_lazy('"Efficiency is doing better what is already being done." — Peter Drucker'),
    gettext_lazy('"The best way to predict the future is to create it." — Peter Drucker'),
]


def _time_based_greeting():
    hour = timezone.localtime().hour

    if hour < 12:
        return gettext("Good morning")
    if hour < 18:
        return gettext("Good afternoon")
    return gettext("Good evening")


def _book_export_rows(books):
    for book in books:
        yield [
            book.isbn or "",
            _csv_safe(book.title),
            _csv_safe(book.subtitle),
            _csv_safe(", ".join(author.name for author in book.authors.all())),
            _csv_safe(book.publisher),
            book.published_date.isoformat(),
            _csv_safe(book.category.name),
            book.distribution_expense,
        ]


def _sale_export_headers():
    return [
        gettext("Date"),
        gettext("Book"),
        gettext("Category"),
        gettext("Quantity"),
        gettext("Unit Price"),
        gettext("Currency"),
        gettext("Tax Rate (%)"),
        gettext("Tax Amount"),
        gettext("Total"),
        gettext("Channel"),
    ]


_PDF_FONTS_REGISTERED = False


def _register_pdf_fonts():
    global _PDF_FONTS_REGISTERED

    if _PDF_FONTS_REGISTERED:
        return

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Bundled (OFL-licensed) so Arabic PDFs render correctly on any host, not
    # just Windows. No separate bold instance is shipped, so both names point
    # at the same regular-weight font.
    font_path = settings.BASE_DIR / "books" / "static" / "books" / "fonts" / "NotoSansArabic.ttf"
    pdfmetrics.registerFont(TTFont("NotoSansArabic", str(font_path)))
    pdfmetrics.registerFont(TTFont("NotoSansArabic-Bold", str(font_path)))

    _PDF_FONTS_REGISTERED = True


def _pdf_fonts():
    if translation.get_language() == "ar":
        return "NotoSansArabic", "NotoSansArabic-Bold"

    return "Helvetica", "Helvetica-Bold"


def _pdf_text(value):
    text = str(value)

    if translation.get_language() == "ar":
        import arabic_reshaper
        from bidi.algorithm import get_display

        return get_display(arabic_reshaper.reshape(text))

    return text


def _sale_export_rows(sales):
    for sale in sales:
        yield [
            sale.sale_date.isoformat(),
            _csv_safe(sale.book.title),
            _csv_safe(sale.book.category.name),
            sale.quantity,
            sale.unit_price,
            sale.currency,
            sale.tax_rate,
            sale.tax_amount,
            sale.total,
            _csv_safe(sale.channel),
        ]


def _invoice_export_headers():
    return [
        gettext("Number"),
        gettext("Customer"),
        gettext("Date"),
        gettext("Due"),
        gettext("Currency"),
        gettext("Subtotal"),
        gettext("Tax"),
        gettext("Total"),
        gettext("Status"),
    ]


def _invoice_export_rows(invoices):
    for invoice in invoices:
        yield [
            _csv_safe(invoice.invoice_number),
            _csv_safe(invoice.customer_name),
            invoice.invoice_date.isoformat(),
            invoice.due_date.isoformat() if invoice.due_date else "",
            invoice.currency,
            invoice.subtotal,
            invoice.tax_total,
            invoice.grand_total,
            invoice.get_status_display(),
        ]


def _reorder_export_headers():
    return [
        gettext("Date"),
        gettext("Book"),
        gettext("Supplier"),
        gettext("Quantity"),
        gettext("Unit Cost"),
        gettext("Total Cost"),
        gettext("Status"),
        gettext("Note"),
        gettext("Received"),
    ]


def _reorder_export_rows(reorders):
    for reorder in reorders:
        yield [
            reorder.created_at.date().isoformat(),
            _csv_safe(reorder.book.title),
            _csv_safe(reorder.supplier.name) if reorder.supplier else "",
            reorder.quantity,
            reorder.unit_cost,
            reorder.total_cost,
            reorder.get_status_display(),
            _csv_safe(reorder.note),
            reorder.received_at.date().isoformat() if reorder.received_at else "",
        ]


def _book_filters(request):
    books = Book.objects.filter(account=request.account).select_related("category").prefetch_related("authors")

    search = request.GET.get("q", "").strip()
    category = request.GET.get("category", "")
    author = request.GET.get("author", "")
    publisher = request.GET.get("publisher", "").strip()
    year = request.GET.get("year", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    low_stock = request.GET.get("low_stock", "")

    if search:
        books = books.filter(
            Q(title__icontains=search)
            | Q(subtitle__icontains=search)
            | Q(authors__name__icontains=search)
            | Q(publisher__icontains=search)
            | Q(isbn__icontains=search)
        ).distinct()

    if category and category.isdigit():
        books = books.filter(category_id=category)

    if author and author.isdigit():
        books = books.filter(authors__id=author).distinct()

    if publisher:
        books = books.filter(publisher=publisher)

    if year:
        books = books.filter(published_date__year=year)

    if start_date:
        books = books.filter(published_date__gte=start_date)

    if end_date:
        books = books.filter(published_date__lte=end_date)

    if low_stock == "1":
        books = books.filter(stock_on_hand__lte=F("reorder_threshold"))

    return books


def _filtered_books_for_export(request):
    return _book_filters(request).order_by("title")


def _active_book_filters(request, context):
    query_params = request.GET.copy()
    query_params.pop("page", None)

    def remove(*keys):
        params = query_params.copy()
        for key in keys:
            params.pop(key, None)
        return params.urlencode()

    filters = []

    search = request.GET.get("q", "").strip()
    if search:
        filters.append({"label": gettext('Search: "%(q)s"') % {"q": search}, "url": remove("q")})

    category_id = request.GET.get("category", "")
    if category_id.isdigit():
        category = context["categories"].filter(id=category_id).first()
        if category:
            filters.append({"label": gettext("Category: %(name)s") % {"name": category.name}, "url": remove("category")})

    author_id = request.GET.get("author", "")
    if author_id.isdigit():
        author = context["authors"].filter(id=author_id).first()
        if author:
            filters.append({"label": gettext("Author: %(name)s") % {"name": author.name}, "url": remove("author")})

    publisher = request.GET.get("publisher", "").strip()
    if publisher:
        filters.append({"label": gettext("Publisher: %(name)s") % {"name": publisher}, "url": remove("publisher")})

    if request.GET.get("low_stock") == "1":
        filters.append({"label": gettext("Low stock only"), "url": remove("low_stock")})

    return filters


def _filter_context(request):
    query_params = request.GET.copy()

    if "page" in query_params:
        query_params.pop("page")

    return {
        "categories": Category.objects.filter(account=request.account).order_by("name"),
        "authors": Author.objects.filter(account=request.account).order_by("name"),
        "publishers": (
            Book.objects.filter(account=request.account)
            .exclude(publisher="")
            .order_by("publisher")
            .values_list("publisher", flat=True)
            .distinct()
        ),
        "years": Book.objects.filter(account=request.account).dates("published_date", "year", order="DESC"),
        "filters": request.GET,
        "query_string": query_params.urlencode(),
    }


@login_required
@permission_required("books.delete_book", raise_exception=True)
@require_POST
def book_bulk_delete(request):
    ids = request.POST.getlist("book_ids")
    if ids:
        deleted, _ = Book.objects.filter(id__in=ids, account=request.account).delete()
        messages.success(request, gettext("%(count)s book(s) deleted.") % {"count": deleted})
    else:
        messages.warning(request, gettext("No books selected."))
    return redirect("book_list")


@login_required
@permission_required("books.view_book", raise_exception=True)
def book_list(request):
    books = _book_filters(request)
    sort = request.GET.get("sort", "title")

    allowed_sorts = {
        "title": "title",
        "-title": "-title",
        "category": "category__name",
        "-category": "-category__name",
        "date": "published_date",
        "-date": "-published_date",
        "expense": "distribution_expense",
        "-expense": "-distribution_expense",
    }

    books = books.order_by(allowed_sorts.get(sort, "title"))
    paginator = Paginator(books, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = _filter_context(request)
    context.update(
        {
            "books": page_obj.object_list,
            "page_obj": page_obj,
            "sort": sort,
            "result_count_text": gettext("%(count)s book(s) found") % {"count": paginator.count},
            "has_any_books": Book.objects.filter(account=request.account).exists(),
        }
    )
    context["active_filters"] = _active_book_filters(request, context)

    return render(request, "books/list.html", context)


@login_required
@permission_required("books.view_book", raise_exception=True)
def book_detail(request, id):
    book = get_object_or_404(
        Book.objects.select_related("category").prefetch_related("authors"),
        id=id,
        account=request.account,
    )

    sales_qs = book.sales.filter(account=request.account)
    totals = sales_qs.aggregate(
        total_quantity=Sum("quantity"),
        total_revenue=Sum(REVENUE_EXPRESSION),
    )

    history = []

    for sale in sales_qs:
        history.append({
            "date": sale.sale_date,
            "type": "sale",
            "delta": -sale.quantity,
            "note": f"{sale.quantity} sold" + (f" via {sale.channel}" if sale.channel else ""),
        })

    for ret in Return.objects.filter(account=request.account, sale__book=book).select_related("sale"):
        history.append({
            "date": ret.return_date,
            "type": "return",
            "delta": ret.quantity,
            "note": ret.reason or "-",
        })

    for adj in book.stock_adjustments.filter(account=request.account):
        history.append({
            "date": adj.created_at.date(),
            "type": "adjustment",
            "delta": adj.change,
            "note": (f"{adj.get_reason_display()}: {adj.note}" if adj.note else adj.get_reason_display()),
        })

    history.sort(key=lambda x: x["date"], reverse=True)

    context = {
        "book": book,
        "total_quantity_sold": totals["total_quantity"] or 0,
        "total_revenue": totals["total_revenue"] or 0,
        "history": history[:50],
    }

    return render(request, "books/detail.html", context)


@login_required
@permission_required("books.add_book", raise_exception=True)
def book_create(request):
    form = BookForm(account=request.account)

    if request.method == "POST":
        form = BookForm(request.POST, account=request.account)

        if form.is_valid():
            book = form.save(commit=False)
            book.owner = request.user
            book.account = request.account
            book.save()
            form.save_m2m()
            messages.success(request, gettext("Book created."))
            return redirect("book_list")

    return render(request, "books/form.html", {"form": form})


@login_required
@permission_required("books.change_book", raise_exception=True)
def book_update(request, id):
    book = get_object_or_404(Book, id=id, account=request.account)
    form = BookForm(instance=book, account=request.account)

    if request.method == "POST":
        form = BookForm(request.POST, instance=book, account=request.account)

        if form.is_valid():
            form.save()
            messages.success(request, gettext("Book updated."))
            return redirect("book_list")

    return render(request, "books/form.html", {"form": form})


@login_required
@permission_required("books.delete_book", raise_exception=True)
def book_delete(request, id):
    book = get_object_or_404(Book, id=id, account=request.account)

    if request.method == "POST":
        book.delete()
        messages.success(request, gettext("Book deleted."))
        return redirect("book_list")

    return render(
        request,
        "books/confirm_delete.html",
        {
            "object_type": gettext("book"),
            "object_name": book.title,
            "cancel_url": reverse("book_list"),
        },
    )


@login_required
@permission_required("books.view_book", raise_exception=True)
def stock_list(request):
    books = Book.objects.filter(account=request.account).select_related("category").order_by("stock_on_hand", "title")
    low_only = request.GET.get("low") == "1"
    q = request.GET.get("q", "").strip()

    if low_only:
        books = books.filter(stock_on_hand__lte=F("reorder_threshold"))
    if q:
        books = books.filter(Q(title__icontains=q) | Q(category__name__icontains=q))

    books = _annotate_stock_value(books)
    total_stock_value = books.aggregate(total=Sum("stock_value"))["total"] or 0

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    def remove(*keys):
        params = query_params.copy()
        for key in keys:
            params.pop(key, None)
        return params.urlencode()

    active_filters = []
    if q:
        active_filters.append({"label": gettext('Search: "%(q)s"') % {"q": q}, "url": remove("q")})
    if low_only:
        active_filters.append({"label": gettext("Low stock only"), "url": remove("low")})

    toggle_params = query_params.copy()
    if low_only:
        toggle_params.pop("low", None)
    else:
        toggle_params["low"] = "1"
    toggle_low_url = "?" + toggle_params.urlencode()

    paginator = Paginator(books, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/stock_list.html",
        {
            "books": page_obj.object_list,
            "page_obj": page_obj,
            "low_only": low_only,
            "total_stock_value": total_stock_value,
            "q": q,
            "query_string": query_string,
            "active_filters": active_filters,
            "toggle_low_url": toggle_low_url,
            "result_count_text": gettext("%(count)s book(s) found") % {"count": paginator.count},
            "has_any_books": Book.objects.filter(account=request.account).exists(),
        },
    )


@login_required
@permission_required("books.view_stockadjustment", raise_exception=True)
def stock_adjustment_list(request):
    adjustments = StockAdjustment.objects.filter(account=request.account).select_related("book", "book__category")

    book_id = request.GET.get("book", "").strip()
    reason = request.GET.get("reason", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()

    if book_id.isdigit():
        adjustments = adjustments.filter(book_id=book_id)
    if reason in dict(StockAdjustment.REASON_CHOICES):
        adjustments = adjustments.filter(reason=reason)
    if start_date:
        adjustments = adjustments.filter(created_at__date__gte=start_date)
    if end_date:
        adjustments = adjustments.filter(created_at__date__lte=end_date)

    books_qs = Book.objects.filter(account=request.account).order_by("title")

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    def remove(*keys):
        params = query_params.copy()
        for key in keys:
            params.pop(key, None)
        return params.urlencode()

    active_filters = []
    if book_id.isdigit():
        book = books_qs.filter(id=book_id).first()
        if book:
            active_filters.append({"label": gettext("Book: %(title)s") % {"title": book.title}, "url": remove("book")})
    if reason in dict(StockAdjustment.REASON_CHOICES):
        active_filters.append({"label": dict(StockAdjustment.REASON_CHOICES)[reason], "url": remove("reason")})
    if start_date or end_date:
        active_filters.append({
            "label": gettext("Date: %(start)s – %(end)s") % {
                "start": start_date or gettext("any"),
                "end": end_date or gettext("any"),
            },
            "url": remove("start_date", "end_date"),
        })

    paginator = Paginator(adjustments, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/stock_adjustment_list.html",
        {
            "adjustments": page_obj.object_list,
            "page_obj": page_obj,
            "books": books_qs,
            "reason_choices": StockAdjustment.REASON_CHOICES,
            "filters": request.GET,
            "query_string": query_string,
            "active_filters": active_filters,
            "result_count_text": gettext("%(count)s adjustment(s) found") % {"count": paginator.count},
            "has_any_adjustments": StockAdjustment.objects.filter(account=request.account).exists(),
        },
    )


@login_required
@permission_required("books.add_stockadjustment", raise_exception=True)
def stock_adjustment_create(request, book_id):
    book = get_object_or_404(Book, id=book_id, account=request.account)

    if request.method == "POST":
        form = StockAdjustmentForm(request.POST)

        if form.is_valid():
            change = form.cleaned_data["change"]

            if book.stock_on_hand + change < 0:
                form.add_error(
                    "change",
                    gettext("This would reduce stock below zero (current stock: %(stock)s).")
                    % {"stock": book.stock_on_hand},
                )
            else:
                adjustment = form.save(commit=False)
                adjustment.owner = request.user
                adjustment.account = request.account
                adjustment.book = book

                book = _adjust_stock(book.id, change, request.user, request.account)
                adjustment.resulting_stock = book.stock_on_hand
                adjustment.save()

                messages.success(request, gettext("Stock adjustment recorded."))
                _notify_stock_level(request, book)
                return redirect("stock_list")
    else:
        form = StockAdjustmentForm()

    return render(
        request,
        "books/stock_adjustment_form.html",
        {
            "form": form,
            "book": book,
        },
    )


@login_required
@permission_required("books.view_category", raise_exception=True)
def category_list(request):
    categories = Category.objects.filter(account=request.account).annotate(book_count=Count("book")).order_by("name")
    q = request.GET.get("q", "").strip()
    if q:
        categories = categories.filter(name__icontains=q)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    paginator = Paginator(categories, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/category_list.html",
        {
            "categories": page_obj.object_list,
            "page_obj": page_obj,
            "q": q,
            "query_string": query_string,
            "result_count_text": gettext("%(count)s category(ies) found") % {"count": paginator.count},
            "has_any_categories": Category.objects.filter(account=request.account).exists(),
        },
    )


@login_required
@permission_required("books.add_category", raise_exception=True)
def category_create(request):
    form = CategoryForm()

    if request.method == "POST":
        form = CategoryForm(request.POST)

        if form.is_valid():
            category = form.save(commit=False)
            category.owner = request.user
            category.account = request.account
            category.save()
            messages.success(request, gettext("Category created."))
            return redirect("category_list")

    return render(request, "books/category_form.html", {"form": form})


@login_required
@permission_required("books.change_category", raise_exception=True)
def category_update(request, id):
    category = get_object_or_404(Category, id=id, account=request.account)
    form = CategoryForm(instance=category)

    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)

        if form.is_valid():
            form.save()
            messages.success(request, gettext("Category updated."))
            return redirect("category_list")

    return render(
        request,
        "books/category_form.html",
        {
            "form": form,
            "category": category,
        },
    )


@login_required
@permission_required("books.delete_category", raise_exception=True)
def category_delete(request, id):
    category = get_object_or_404(Category, id=id, account=request.account)
    book_count = category.book_set.count()

    if request.method == "POST":
        if book_count:
            messages.error(
                request,
                gettext("Move or delete this category's books before deleting the category."),
            )
            return redirect("category_list")

        category.delete()
        messages.success(request, gettext("Category deleted."))
        return redirect("category_list")

    return render(
        request,
        "books/confirm_delete.html",
        {
            "object_type": gettext("category"),
            "object_name": category.name,
            "cancel_url": reverse("category_list"),
            "warning": (
                gettext("This category contains books and cannot be deleted yet.")
                if book_count
                else ""
            ),
            "disable_delete": bool(book_count),
        },
    )


@login_required
@permission_required("books.view_author", raise_exception=True)
def author_list(request):
    authors = Author.objects.filter(account=request.account).annotate(book_count=Count("books")).order_by("name")
    q = request.GET.get("q", "").strip()
    if q:
        authors = authors.filter(name__icontains=q)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    paginator = Paginator(authors, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/author_list.html",
        {
            "authors": page_obj.object_list,
            "page_obj": page_obj,
            "q": q,
            "query_string": query_string,
            "result_count_text": gettext("%(count)s author(s) found") % {"count": paginator.count},
            "has_any_authors": Author.objects.filter(account=request.account).exists(),
        },
    )


@login_required
@permission_required("books.add_author", raise_exception=True)
def author_create(request):
    form = AuthorForm()

    if request.method == "POST":
        form = AuthorForm(request.POST)

        if form.is_valid():
            author = form.save(commit=False)
            author.owner = request.user
            author.account = request.account
            author.save()
            messages.success(request, gettext("Author created."))
            return redirect("author_list")

    return render(request, "books/author_form.html", {"form": form})


@login_required
@permission_required("books.change_author", raise_exception=True)
def author_update(request, id):
    author = get_object_or_404(Author, id=id, account=request.account)
    form = AuthorForm(instance=author)

    if request.method == "POST":
        form = AuthorForm(request.POST, instance=author)

        if form.is_valid():
            form.save()
            messages.success(request, gettext("Author updated."))
            return redirect("author_list")

    return render(
        request,
        "books/author_form.html",
        {
            "form": form,
            "author": author,
        },
    )


@login_required
@permission_required("books.delete_author", raise_exception=True)
def author_delete(request, id):
    author = get_object_or_404(Author, id=id, account=request.account)
    book_count = author.books.count()

    if request.method == "POST":
        if book_count:
            messages.error(
                request,
                gettext("Remove this author from their books before deleting them."),
            )
            return redirect("author_list")

        author.delete()
        messages.success(request, gettext("Author deleted."))
        return redirect("author_list")

    return render(
        request,
        "books/confirm_delete.html",
        {
            "object_type": gettext("author"),
            "object_name": author.name,
            "cancel_url": reverse("author_list"),
            "warning": (
                gettext("This author is linked to books and cannot be deleted yet.")
                if book_count
                else ""
            ),
            "disable_delete": bool(book_count),
        },
    )


@login_required
@permission_required("books.view_supplier", raise_exception=True)
def supplier_list(request):
    suppliers = Supplier.objects.filter(account=request.account).annotate(reorder_count=Count("reorders")).order_by("name")
    q = request.GET.get("q", "").strip()
    if q:
        suppliers = suppliers.filter(
            Q(name__icontains=q) | Q(contact_name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q)
        )

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    paginator = Paginator(suppliers, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/supplier_list.html",
        {
            "suppliers": page_obj.object_list,
            "page_obj": page_obj,
            "q": q,
            "query_string": query_string,
            "result_count_text": gettext("%(count)s supplier(s) found") % {"count": paginator.count},
            "has_any_suppliers": Supplier.objects.filter(account=request.account).exists(),
        },
    )


@login_required
@permission_required("books.add_supplier", raise_exception=True)
def supplier_create(request):
    form = SupplierForm()

    if request.method == "POST":
        form = SupplierForm(request.POST)

        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.owner = request.user
            supplier.account = request.account
            supplier.save()
            messages.success(request, gettext("Supplier created."))
            return redirect("supplier_list")

    return render(request, "books/supplier_form.html", {"form": form})


@login_required
@permission_required("books.change_supplier", raise_exception=True)
def supplier_update(request, id):
    supplier = get_object_or_404(Supplier, id=id, account=request.account)
    form = SupplierForm(instance=supplier)

    if request.method == "POST":
        form = SupplierForm(request.POST, instance=supplier)

        if form.is_valid():
            form.save()
            messages.success(request, gettext("Supplier updated."))
            return redirect("supplier_list")

    return render(
        request,
        "books/supplier_form.html",
        {
            "form": form,
            "supplier": supplier,
        },
    )


@login_required
@permission_required("books.delete_supplier", raise_exception=True)
def supplier_delete(request, id):
    supplier = get_object_or_404(Supplier, id=id, account=request.account)
    reorder_count = supplier.reorders.count()

    if request.method == "POST":
        if reorder_count:
            messages.error(
                request,
                gettext("Remove this supplier from its reorders before deleting it."),
            )
            return redirect("supplier_list")

        supplier.delete()
        messages.success(request, gettext("Supplier deleted."))
        return redirect("supplier_list")

    return render(
        request,
        "books/confirm_delete.html",
        {
            "object_type": gettext("supplier"),
            "object_name": supplier.name,
            "cancel_url": reverse("supplier_list"),
            "warning": (
                gettext("This supplier is linked to reorders and cannot be deleted yet.")
                if reorder_count
                else ""
            ),
            "disable_delete": bool(reorder_count),
        },
    )


SALE_TOTAL_EXPRESSION = ExpressionWrapper(
    F("quantity") * F("unit_price") * (1 + F("tax_rate") / Decimal("100")),
    output_field=DecimalField(max_digits=10, decimal_places=2),
)

_RETURN_AMOUNT_EXPRESSION = ExpressionWrapper(
    F("quantity") * F("sale__unit_price"),
    output_field=DecimalField(max_digits=10, decimal_places=2),
)


def _annotate_stock_value(books):
    latest_cost = (
        Reorder.objects.filter(
            book=OuterRef("pk"),
            status=Reorder.STATUS_RECEIVED,
            received_at__isnull=False,
        )
        .order_by("-received_at")
        .values("unit_cost")[:1]
    )

    return books.annotate(
        unit_cost=Coalesce(
            Subquery(latest_cost, output_field=DecimalField(max_digits=8, decimal_places=2)),
            Value(0, output_field=DecimalField(max_digits=8, decimal_places=2)),
        ),
    ).annotate(
        stock_value=ExpressionWrapper(
            F("stock_on_hand") * F("unit_cost"),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )


def _safe_json(value):
    """json.dumps for embedding in a <script> block via the |safe filter.

    Escapes "</" so user-entered strings (category names, etc.) can't contain
    a literal "</script>" that would terminate the tag early and inject HTML.
    """
    return json.dumps(value).replace("</", "<\\/")


@login_required
@permission_required("books.view_book", raise_exception=True)
def report(request):
    filtered_books = _book_filters(request)

    data = (
        filtered_books
        .values("category__name")
        .annotate(total=Sum("distribution_expense"), count=Count("id"))
        .order_by("category__name")
    )

    revenue_by_category = {
        item["book__category__name"]: item["revenue"] or 0
        for item in (
            Sale.objects.filter(book__in=filtered_books)
            .values("book__category__name")
            .annotate(revenue=Sum(REVENUE_EXPRESSION))
        )
    }

    received_reorders = Reorder.objects.filter(
        book__in=filtered_books, status=Reorder.STATUS_RECEIVED
    )

    purchase_cost_by_category = {
        item["book__category__name"]: item["cost"] or 0
        for item in (
            received_reorders
            .values("book__category__name")
            .annotate(cost=Sum(PURCHASE_COST_EXPRESSION))
        )
    }

    labels = []
    values = []
    counts = []
    revenues = []
    profits = []
    purchase_costs = []

    for item in data:
        category_name = item["category__name"]
        expense = float(item["total"])
        revenue = float(revenue_by_category.get(category_name, 0))
        purchase_cost = float(purchase_cost_by_category.get(category_name, 0))

        labels.append(category_name)
        values.append(expense)
        counts.append(item["count"])
        revenues.append(revenue)
        profits.append(revenue - expense)
        purchase_costs.append(purchase_cost)

    totals = filtered_books.aggregate(
        total=Sum("distribution_expense"),
        count=Count("id"),
    )

    total_revenue = (
        Sale.objects.filter(book__in=filtered_books).aggregate(
            total=Sum(REVENUE_EXPRESSION)
        )["total"]
        or 0
    )
    total_expense = totals["total"] or 0
    total_purchase_cost = (
        received_reorders.aggregate(total=Sum(PURCHASE_COST_EXPRESSION))["total"] or 0
    )

    sales_trend = (
        Sale.objects.filter(book__in=filtered_books)
        .annotate(month=TruncMonth("sale_date"))
        .values("month")
        .annotate(units=Sum("quantity"), revenue=Sum(REVENUE_EXPRESSION))
    )

    purchase_trend = (
        received_reorders.filter(received_at__isnull=False)
        .annotate(month=TruncMonth("received_at"))
        .values("month")
        .annotate(cost=Sum(PURCHASE_COST_EXPRESSION))
    )

    sales_by_month = {item["month"]: item for item in sales_trend}
    purchase_by_month = {
        item["month"].date(): item for item in purchase_trend
    }

    trend_labels = []
    trend_units = []
    trend_revenues = []
    trend_purchase_costs = []

    for month in sorted(set(sales_by_month) | set(purchase_by_month)):
        trend_labels.append(month.strftime("%b %Y"))
        trend_units.append(sales_by_month.get(month, {}).get("units") or 0)
        trend_revenues.append(float(sales_by_month.get(month, {}).get("revenue") or 0))
        trend_purchase_costs.append(float(purchase_by_month.get(month, {}).get("cost") or 0))

    context = _filter_context(request)
    context.update(
        {
            "values": _safe_json(values),
            "counts": _safe_json(counts),
            "labels": _safe_json(labels),
            "revenues": _safe_json(revenues),
            "profits": _safe_json(profits),
            "purchase_costs": _safe_json(purchase_costs),
            "trend_labels": _safe_json(trend_labels),
            "trend_units": _safe_json(trend_units),
            "trend_revenues": _safe_json(trend_revenues),
            "trend_purchase_costs": _safe_json(trend_purchase_costs),
            "total_expense": total_expense,
            "total_books": totals["count"],
            "total_revenue": total_revenue,
            "total_profit": total_revenue - total_expense,
            "total_purchase_cost": total_purchase_cost,
        }
    )

    return render(request, "books/report.html", context)


@login_required
@permission_required("books.view_book", raise_exception=True)
def dashboard(request):
    books = Book.objects.filter(account=request.account).select_related("category")
    sales = Sale.objects.filter(account=request.account)

    total_expense = books.aggregate(total=Sum("distribution_expense"))["total"] or 0
    total_revenue = sales.aggregate(total=Sum(REVENUE_EXPRESSION))["total"] or 0
    total_profit = total_revenue - total_expense
    total_units_sold = sales.aggregate(total=Sum("quantity"))["total"] or 0

    low_stock_books = books.filter(
        stock_on_hand__lte=F("reorder_threshold")
    ).order_by("stock_on_hand", "title")

    category_expenses = (
        books.values("category__name")
        .annotate(expense=Sum("distribution_expense"))
        .order_by("category__name")
    )

    revenue_by_category = {
        item["book__category__name"]: item["revenue"] or 0
        for item in (
            sales.values("book__category__name")
            .annotate(revenue=Sum(REVENUE_EXPRESSION))
        )
    }

    received_reorders = Reorder.objects.filter(account=request.account, status=Reorder.STATUS_RECEIVED)

    purchase_cost_by_category = {
        item["book__category__name"]: item["cost"] or 0
        for item in (
            received_reorders.values("book__category__name")
            .annotate(cost=Sum(PURCHASE_COST_EXPRESSION))
        )
    }

    labels = []
    revenues = []
    profits = []
    purchase_costs = []

    for item in category_expenses:
        category_name = item["category__name"]
        expense = float(item["expense"] or 0)
        revenue = float(revenue_by_category.get(category_name, 0))
        purchase_cost = float(purchase_cost_by_category.get(category_name, 0))

        labels.append(category_name)
        revenues.append(revenue)
        profits.append(revenue - expense)
        purchase_costs.append(purchase_cost)

    total_purchase_cost = (
        received_reorders.aggregate(total=Sum(PURCHASE_COST_EXPRESSION))["total"] or 0
    )

    total_inventory_value = (
        _annotate_stock_value(books).aggregate(total=Sum("stock_value"))["total"] or 0
    )

    sales_trend = (
        sales.annotate(month=TruncMonth("sale_date"))
        .values("month")
        .annotate(units=Sum("quantity"), revenue=Sum(REVENUE_EXPRESSION))
    )

    purchase_trend = (
        received_reorders.filter(received_at__isnull=False)
        .annotate(month=TruncMonth("received_at"))
        .values("month")
        .annotate(cost=Sum(PURCHASE_COST_EXPRESSION))
    )

    sales_by_month = {item["month"]: item for item in sales_trend}
    purchase_by_month = {item["month"].date(): item for item in purchase_trend}

    trend_labels = []
    trend_units = []
    trend_revenues = []
    trend_purchase_costs = []

    for month in sorted(set(sales_by_month) | set(purchase_by_month)):
        trend_labels.append(month.strftime("%b %Y"))
        trend_units.append(sales_by_month.get(month, {}).get("units") or 0)
        trend_revenues.append(float(sales_by_month.get(month, {}).get("revenue") or 0))
        trend_purchase_costs.append(float(purchase_by_month.get(month, {}).get("cost") or 0))

    channel_breakdown = list(
        sales.annotate(channel_name=Coalesce(
            Case(
                When(channel="", then=Value(None)),
                default=F("channel"),
                output_field=models.CharField(),
            ),
            Value(gettext("Unspecified")),
        ))
        .values("channel_name")
        .annotate(revenue=Sum(REVENUE_EXPRESSION), units=Sum("quantity"))
        .order_by("-revenue")
    )

    channel_labels = [item["channel_name"] for item in channel_breakdown]
    channel_revenues = [float(item["revenue"] or 0) for item in channel_breakdown]
    channel_total_revenue = sum(channel_revenues) or 1
    channel_rows = [
        {
            "channel": item["channel_name"],
            "revenue": item["revenue"] or 0,
            "units": item["units"] or 0,
            "share": round(float(item["revenue"] or 0) / channel_total_revenue * 100, 1),
        }
        for item in channel_breakdown
    ]

    top_books = (
        books.annotate(units_sold=Sum("sales__quantity"))
        .filter(units_sold__gt=0)
        .order_by("-units_sold")[:5]
    )

    recent_sales = sales.select_related("book", "book__category")[:5]

    pending_reorders_count = Reorder.objects.filter(
        account=request.account,
        status__in=[Reorder.STATUS_PENDING, Reorder.STATUS_ORDERED],
    ).count()

    today = timezone.now().date()
    overdue_invoices_count = Invoice.objects.filter(
        account=request.account,
        due_date__lt=today,
        due_date__isnull=False,
    ).exclude(status=Invoice.STATUS_PAID).count()

    low_stock_list = list(low_stock_books[:5])
    velocity_cutoff = timezone.now().date() - timedelta(days=REORDER_VELOCITY_WINDOW_DAYS)
    velocity_by_book = {
        item["book_id"]: (item["units"] or 0) / REORDER_VELOCITY_WINDOW_DAYS
        for item in Sale.objects.filter(
            book_id__in=[b.id for b in low_stock_list],
            sale_date__gte=velocity_cutoff,
        ).values("book_id").annotate(units=Sum("quantity"))
    }

    context = {
        "greeting": _time_based_greeting(),
        "quote": random.choice(LEARNING_QUOTES),
        "all_quotes": [str(quote) for quote in LEARNING_QUOTES],
        "total_books": books.count(),
        "total_authors": Author.objects.filter(account=request.account).count(),
        "total_categories": Category.objects.filter(account=request.account).count(),
        "low_stock_count": low_stock_books.count(),
        "low_stock_books": [
            {
                "book": book,
                "suggested_quantity": _suggested_reorder_quantity(
                    book, velocity=velocity_by_book.get(book.id, 0)
                ),
            }
            for book in low_stock_list
        ],
        "total_revenue": total_revenue,
        "total_expense": total_expense,
        "total_profit": total_profit,
        "total_units_sold": total_units_sold,
        "total_purchase_cost": total_purchase_cost,
        "total_inventory_value": total_inventory_value,
        "labels": _safe_json(labels),
        "revenues": _safe_json(revenues),
        "profits": _safe_json(profits),
        "purchase_costs": _safe_json(purchase_costs),
        "trend_labels": _safe_json(trend_labels),
        "trend_units": _safe_json(trend_units),
        "trend_revenues": _safe_json(trend_revenues),
        "trend_purchase_costs": _safe_json(trend_purchase_costs),
        "channel_labels": _safe_json(channel_labels),
        "channel_revenues": _safe_json(channel_revenues),
        "channel_rows": channel_rows,
        "top_books": top_books,
        "recent_sales": recent_sales,
        "pending_reorders_count": pending_reorders_count,
        "overdue_invoices_count": overdue_invoices_count,
    }

    return render(request, "books/dashboard.html", context)


def _bootstrap_widgets(form):
    for field in form.fields.values():
        field.widget.attrs.setdefault("class", "form-control")
    return form


@login_required
def profile_update(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    avatar_form = ProfileForm(instance=profile)
    email_form = _bootstrap_widgets(EmailUpdateForm(initial={"email": request.user.email}))
    password_form = _bootstrap_widgets(PasswordChangeForm(user=request.user))

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "avatar":
            avatar_form = ProfileForm(request.POST, request.FILES, instance=profile)
            if avatar_form.is_valid():
                avatar_form.save()
                messages.success(request, gettext("Profile photo updated."))
                return redirect("profile_update")

        elif action == "email":
            email_form = _bootstrap_widgets(EmailUpdateForm(request.POST))
            if email_form.is_valid():
                request.user.email = email_form.cleaned_data["email"]
                request.user.save(update_fields=["email"])
                messages.success(request, gettext("Email address updated."))
                return redirect("profile_update")

        elif action == "password":
            password_form = _bootstrap_widgets(PasswordChangeForm(request.user, request.POST))
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, gettext("Password updated."))
                return redirect("profile_update")

    return render(request, "books/profile_form.html", {
        "avatar_form": avatar_form,
        "email_form": email_form,
        "password_form": password_form,
        "profile": profile,
    })


SEARCH_RESULT_LIMIT = 25


@login_required
def global_search(request):
    query = request.GET.get("q", "").strip()

    books = customers = invoices = []

    if query:
        if request.user.has_perm("books.view_book"):
            books = Book.objects.filter(account=request.account).filter(
                Q(title__icontains=query) | Q(isbn__icontains=query) | Q(publisher__icontains=query)
            ).select_related("category")[:SEARCH_RESULT_LIMIT]

        if request.user.has_perm("books.view_customer"):
            customers = Customer.objects.filter(account=request.account).filter(
                Q(name__icontains=query) | Q(email__icontains=query) | Q(phone__icontains=query)
            )[:SEARCH_RESULT_LIMIT]

        if request.user.has_perm("books.view_invoice"):
            invoices = Invoice.objects.filter(account=request.account).filter(
                Q(invoice_number__icontains=query) | Q(customer_name__icontains=query)
            )[:SEARCH_RESULT_LIMIT]

    return render(request, "books/global_search.html", {
        "query": query,
        "books": books,
        "customers": customers,
        "invoices": invoices,
        "has_results": bool(books or customers or invoices),
    })


@login_required
def about(request):
    return render(request, "books/about.html")


@login_required
@require_POST
def chat_api(request):
    try:
        payload = json.loads(request.body)
    except ValueError:
        return JsonResponse({"error": "Invalid JSON."}, status=400)

    message = (payload.get("message") or "").strip()
    history = payload.get("history") or []

    if not message:
        return JsonResponse({"error": "Message is required."}, status=400)

    cooldown_key = f"chat_api_cooldown:{request.user.id}"
    if not cache.add(cooldown_key, True, timeout=settings.CHAT_API_COOLDOWN_SECONDS):
        return JsonResponse({"error": gettext("Please wait a moment before sending another message.")}, status=429)

    try:
        reply, updated_history = ai_chat.get_chat_reply(request.user, request.account, message, history)
    except Exception:
        logger.exception("AI chat request failed")
        reply = gettext("Sorry, something went wrong talking to the AI assistant. Please try again.")
        updated_history = history

    return JsonResponse({"reply": reply, "history": updated_history})


def _adjust_stock(book_id, delta, owner, account, location=None):
    # select_for_update() serializes concurrent adjustments to the same book
    # (sales, returns, reorder receipts, transfers, the Shopify webhook all
    # funnel through here) - without it, two requests racing past the
    # read-then-write below can silently drop one of the two updates.
    with transaction.atomic():
        book = Book.objects.select_for_update().get(id=book_id, account=account)

        if location is None:
            location, _ = Location.objects.get_or_create(
                account=account,
                is_default=True,
                defaults={"name": "Main Warehouse", "owner": owner},
            )

        book_has_stock_levels = StockLevel.objects.filter(book=book).exists()

        level, created = StockLevel.objects.select_for_update().get_or_create(
            account=account, book=book, location=location,
            defaults={"quantity": 0, "owner": owner},
        )

        # First-ever StockLevel for this book: seed from stock_on_hand so location
        # tracking starts from reality instead of silently discarding it. Skip this
        # when the book already has other locations, since their total is already
        # reflected in stock_on_hand and seeding again would double-count it.
        if created and not book_has_stock_levels:
            level.quantity = book.stock_on_hand

        level.quantity = max(0, level.quantity + delta)
        level.save(update_fields=["quantity"])

        total = StockLevel.objects.filter(book=book).aggregate(t=Sum("quantity"))["t"] or 0
        book.stock_on_hand = total

        if book.is_low_stock and not book.low_stock_alert_sent:
            book.low_stock_alert_sent = True
        elif not book.is_low_stock and book.low_stock_alert_sent:
            book.low_stock_alert_sent = False

        book.save(update_fields=["stock_on_hand", "low_stock_alert_sent"])
    return book


def _send_reorder_status_email(user, reorder):
    if not user.email:
        return

    status_label = reorder.get_status_display()

    send_mail(
        subject=f"RumiPress: Reorder {status_label} - {reorder.book.title}",
        message=(
            f"Hi {user.username},\n\n"
            f"Your reorder for '{reorder.book.title}' (quantity {reorder.quantity}) "
            f"is now marked as {status_label.lower()}."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=True,
    )


def _notify_stock_level(request, book):
    if book.is_low_stock:
        messages.warning(
            request,
            gettext(
                "Low stock: '%(title)s' has %(stock)s remaining "
                "(reorder threshold %(threshold)s)."
            )
            % {
                "title": book.title,
                "stock": book.stock_on_hand,
                "threshold": book.reorder_threshold,
            },
        )
    else:
        messages.info(
            request,
            gettext("'%(title)s' now has %(stock)s in stock.")
            % {"title": book.title, "stock": book.stock_on_hand},
        )


def _sale_filters(request):
    sales = Sale.objects.filter(account=request.account).select_related("book", "book__category")
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    book_id = request.GET.get("book", "")
    channel = request.GET.get("channel", "").strip()

    if start_date:
        sales = sales.filter(sale_date__gte=start_date)
    if end_date:
        sales = sales.filter(sale_date__lte=end_date)
    if book_id and book_id.isdigit():
        sales = sales.filter(book_id=book_id)
    if channel:
        sales = sales.filter(channel=channel)

    return sales.order_by("-sale_date")


def _active_sale_filters(request, books_qs):
    query_params = request.GET.copy()
    query_params.pop("page", None)

    def remove(*keys):
        params = query_params.copy()
        for key in keys:
            params.pop(key, None)
        return params.urlencode()

    filters = []

    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    if start_date or end_date:
        filters.append({
            "label": gettext("Date: %(start)s – %(end)s") % {
                "start": start_date or gettext("any"),
                "end": end_date or gettext("any"),
            },
            "url": remove("start_date", "end_date"),
        })

    book_id = request.GET.get("book", "")
    if book_id.isdigit():
        book = books_qs.filter(id=book_id).first()
        if book:
            filters.append({"label": gettext("Book: %(title)s") % {"title": book.title}, "url": remove("book")})

    channel = request.GET.get("channel", "").strip()
    if channel:
        filters.append({"label": gettext("Channel: %(channel)s") % {"channel": channel}, "url": remove("channel")})

    return filters


@login_required
@permission_required("books.view_sale", raise_exception=True)
def sale_list(request):
    sales = _sale_filters(request)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    books_qs = Book.objects.filter(account=request.account).order_by("title")
    channels = (
        Sale.objects.filter(account=request.account)
        .exclude(channel="")
        .values_list("channel", flat=True)
        .distinct()
        .order_by("channel")
    )

    totals_by_currency = list(
        sales.values("currency")
        .annotate(total_quantity=Sum("quantity"), total_amount=Sum(SALE_TOTAL_EXPRESSION))
        .order_by("currency")
    )

    paginator = Paginator(sales, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/sale_list.html",
        {
            "sales": page_obj.object_list,
            "page_obj": page_obj,
            "books": books_qs,
            "channels": channels,
            "filters": request.GET,
            "query_string": query_string,
            "active_filters": _active_sale_filters(request, books_qs),
            "result_count_text": gettext("%(count)s sale(s) found") % {"count": paginator.count},
            "has_any_sales": Sale.objects.filter(account=request.account).exists(),
            "totals_by_currency": totals_by_currency,
        },
    )


@login_required
@permission_required("books.add_sale", raise_exception=True)
def sale_create(request):
    form = SaleForm(account=request.account)

    if request.method == "POST":
        form = SaleForm(request.POST, account=request.account)

        if form.is_valid():
            book = form.cleaned_data["book"]
            quantity = form.cleaned_data["quantity"]

            if book.stock_on_hand < quantity:
                messages.error(
                    request,
                    gettext("Cannot record sale: '%(title)s' only has %(stock)s in stock.")
                    % {"title": book.title, "stock": book.stock_on_hand},
                )
            else:
                sale = form.save(commit=False)
                sale.owner = request.user
                sale.account = request.account
                sale.save()
                book = _adjust_stock(sale.book_id, -sale.quantity, request.user, request.account)
                messages.success(request, gettext("Sale recorded."))
                _notify_stock_level(request, book)
                return redirect("sale_list")

    return render(request, "books/sale_form.html", {"form": form})


@login_required
@permission_required("books.change_sale", raise_exception=True)
def sale_update(request, id):
    sale = get_object_or_404(Sale, id=id, account=request.account)
    form = SaleForm(instance=sale, account=request.account)

    if request.method == "POST":
        previous_book_id = sale.book_id
        previous_quantity = sale.quantity

        form = SaleForm(request.POST, instance=sale, account=request.account)

        if form.is_valid():
            new_book = form.cleaned_data["book"]
            new_quantity = form.cleaned_data["quantity"]

            available = new_book.stock_on_hand
            if new_book.id == previous_book_id:
                available += previous_quantity

            if available < new_quantity:
                messages.error(
                    request,
                    gettext("Cannot update sale: '%(title)s' only has %(stock)s available.")
                    % {"title": new_book.title, "stock": available},
                )
            else:
                sale = form.save()
                _adjust_stock(previous_book_id, previous_quantity, request.user, request.account)
                book = _adjust_stock(sale.book_id, -sale.quantity, request.user, request.account)
                messages.success(request, gettext("Sale updated."))
                _notify_stock_level(request, book)
                return redirect("sale_list")

    return render(
        request,
        "books/sale_form.html",
        {
            "form": form,
            "sale": sale,
        },
    )


@login_required
@permission_required("books.delete_sale", raise_exception=True)
def sale_delete(request, id):
    sale = get_object_or_404(Sale, id=id, account=request.account)

    if request.method == "POST":
        _adjust_stock(sale.book_id, sale.quantity, request.user, request.account)
        sale.delete()
        messages.success(request, gettext("Sale deleted."))
        return redirect("sale_list")

    return render(
        request,
        "books/confirm_delete.html",
        {
            "object_type": gettext("sale"),
            "object_name": f"{sale.book.title} ({sale.sale_date})",
            "cancel_url": reverse("sale_list"),
        },
    )


@login_required
@permission_required("books.view_return", raise_exception=True)
def return_list(request):
    returns = Return.objects.filter(account=request.account).select_related("sale", "sale__book")

    book_id = request.GET.get("book", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()

    if book_id.isdigit():
        returns = returns.filter(sale__book_id=book_id)
    if start_date:
        returns = returns.filter(return_date__gte=start_date)
    if end_date:
        returns = returns.filter(return_date__lte=end_date)

    books_qs = Book.objects.filter(account=request.account).order_by("title")

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    def remove(*keys):
        params = query_params.copy()
        for key in keys:
            params.pop(key, None)
        return params.urlencode()

    active_filters = []
    if book_id.isdigit():
        book = books_qs.filter(id=book_id).first()
        if book:
            active_filters.append({"label": gettext("Book: %(title)s") % {"title": book.title}, "url": remove("book")})
    if start_date or end_date:
        active_filters.append({
            "label": gettext("Date: %(start)s – %(end)s") % {
                "start": start_date or gettext("any"),
                "end": end_date or gettext("any"),
            },
            "url": remove("start_date", "end_date"),
        })

    refund_totals_by_currency = list(
        returns.values("sale__currency")
        .annotate(total_amount=Sum(_RETURN_AMOUNT_EXPRESSION))
        .order_by("sale__currency")
    )

    paginator = Paginator(returns, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/return_list.html",
        {
            "returns": page_obj.object_list,
            "page_obj": page_obj,
            "books": books_qs,
            "filters": request.GET,
            "query_string": query_string,
            "active_filters": active_filters,
            "result_count_text": gettext("%(count)s return(s) found") % {"count": paginator.count},
            "has_any_returns": Return.objects.filter(account=request.account).exists(),
            "refund_totals_by_currency": refund_totals_by_currency,
        },
    )


@login_required
@permission_required("books.add_return", raise_exception=True)
def return_create(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id, account=request.account)

    if sale.quantity <= 0:
        messages.error(request, gettext("This sale has already been fully returned."))
        return redirect("sale_list")

    if request.method == "POST":
        form = ReturnForm(request.POST)

        if form.is_valid():
            quantity = form.cleaned_data["quantity"]

            if quantity > sale.quantity:
                messages.error(
                    request,
                    gettext("Cannot return more than the %(quantity)s sold.")
                    % {"quantity": sale.quantity},
                )
            else:
                return_obj = form.save(commit=False)
                return_obj.owner = request.user
                return_obj.account = request.account
                return_obj.sale = sale
                return_obj.save()

                sale.quantity -= quantity
                sale.save(update_fields=["quantity"])

                book = _adjust_stock(sale.book_id, quantity, request.user, request.account)
                messages.success(request, gettext("Return recorded and stock updated."))
                _notify_stock_level(request, book)
                return redirect("return_list")
    else:
        form = ReturnForm(initial={"quantity": sale.quantity, "return_date": timezone.now().date()})

    return render(
        request,
        "books/return_form.html",
        {
            "form": form,
            "sale": sale,
        },
    )


@login_required
@permission_required("books.delete_return", raise_exception=True)
def return_delete(request, id):
    return_obj = get_object_or_404(Return, id=id, account=request.account)

    if request.method == "POST":
        sale = return_obj.sale
        sale.quantity += return_obj.quantity
        sale.save(update_fields=["quantity"])
        _adjust_stock(sale.book_id, -return_obj.quantity, request.user, request.account)
        return_obj.delete()
        messages.success(request, gettext("Return deleted."))
        return redirect("return_list")

    return render(
        request,
        "books/confirm_delete.html",
        {
            "object_type": gettext("return"),
            "object_name": f"{return_obj.sale.book.title} ({return_obj.return_date})",
            "cancel_url": reverse("return_list"),
        },
    )


@login_required
@permission_required("books.view_reorder", raise_exception=True)
def reorder_suggestions(request):
    books = Book.objects.filter(account=request.account).select_related("category")

    cutoff = timezone.now().date() - timedelta(days=REORDER_VELOCITY_WINDOW_DAYS)

    recent_sales = (
        Sale.objects.filter(book=OuterRef("pk"), sale_date__gte=cutoff)
        .values("book")
        .annotate(total=Sum("quantity"))
        .values("total")
    )

    books = books.annotate(
        units_sold_recent=Coalesce(
            Subquery(recent_sales, output_field=IntegerField()),
            Value(0, output_field=IntegerField()),
        )
    )

    suggestions = []

    for book in books:
        velocity = book.units_sold_recent / REORDER_VELOCITY_WINDOW_DAYS
        days_of_stock = book.stock_on_hand / velocity if velocity > 0 else None
        needs_reorder = book.is_low_stock or (
            days_of_stock is not None and days_of_stock <= REORDER_COVER_DAYS
        )

        if not needs_reorder:
            continue

        suggestions.append({
            "book": book,
            "daily_sales_velocity": velocity,
            "days_of_stock": days_of_stock,
            "suggested_quantity": _suggested_reorder_quantity(book, velocity=velocity),
        })

    suggestions.sort(
        key=lambda item: item["days_of_stock"] if item["days_of_stock"] is not None else -1
    )

    return render(
        request,
        "books/reorder_suggestions.html",
        {"suggestions": suggestions},
    )


def _reorder_filters(request):
    reorders = Reorder.objects.filter(account=request.account).select_related("book", "book__category", "supplier")

    status = request.GET.get("status", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    supplier_id = request.GET.get("supplier", "").strip()
    book_id = request.GET.get("book", "").strip()

    if status in dict(Reorder.STATUS_CHOICES):
        reorders = reorders.filter(status=status)
    if start_date:
        reorders = reorders.filter(created_at__date__gte=start_date)
    if end_date:
        reorders = reorders.filter(created_at__date__lte=end_date)
    if supplier_id.isdigit():
        reorders = reorders.filter(supplier_id=supplier_id)
    if book_id.isdigit():
        reorders = reorders.filter(book_id=book_id)

    return reorders


def _active_reorder_filters(request, books_qs, suppliers_qs):
    query_params = request.GET.copy()
    query_params.pop("page", None)

    def remove(*keys):
        params = query_params.copy()
        for key in keys:
            params.pop(key, None)
        return params.urlencode()

    filters = []

    status = request.GET.get("status", "").strip()
    if status in dict(Reorder.STATUS_CHOICES):
        filters.append({"label": dict(Reorder.STATUS_CHOICES)[status], "url": remove("status")})

    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    if start_date or end_date:
        filters.append({
            "label": gettext("Date: %(start)s – %(end)s") % {
                "start": start_date or gettext("any"),
                "end": end_date or gettext("any"),
            },
            "url": remove("start_date", "end_date"),
        })

    supplier_id = request.GET.get("supplier", "").strip()
    if supplier_id.isdigit():
        supplier = suppliers_qs.filter(id=supplier_id).first()
        if supplier:
            filters.append({"label": gettext("Supplier: %(name)s") % {"name": supplier.name}, "url": remove("supplier")})

    book_id = request.GET.get("book", "").strip()
    if book_id.isdigit():
        book = books_qs.filter(id=book_id).first()
        if book:
            filters.append({"label": gettext("Book: %(title)s") % {"title": book.title}, "url": remove("book")})

    return filters


@login_required
@permission_required("books.view_reorder", raise_exception=True)
def reorder_list(request):
    reorders = _reorder_filters(request)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    suppliers = Supplier.objects.filter(account=request.account).order_by("name")
    books_qs = Book.objects.filter(account=request.account).order_by("title")

    paginator = Paginator(reorders, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "books/reorder_list.html",
        {
            "reorders": page_obj.object_list,
            "page_obj": page_obj,
            "status_choices": Reorder.STATUS_CHOICES,
            "selected_status": request.GET.get("status", "").strip(),
            "suppliers": suppliers,
            "books": books_qs,
            "filters": request.GET,
            "query_string": query_string,
            "active_filters": _active_reorder_filters(request, books_qs, suppliers),
            "result_count_text": gettext("%(count)s reorder(s) found") % {"count": paginator.count},
            "has_any_reorders": Reorder.objects.filter(account=request.account).exists(),
        },
    )


@login_required
@permission_required("books.add_reorder", raise_exception=True)
def reorder_create(request, book_id):
    book = get_object_or_404(Book, id=book_id, account=request.account)

    suggested_quantity = _suggested_reorder_quantity(book)

    if request.method == "POST":
        form = ReorderForm(request.POST, account=request.account)

        if form.is_valid():
            reorder = form.save(commit=False)
            reorder.owner = request.user
            reorder.account = request.account
            reorder.book = book
            reorder.save()
            messages.success(request, gettext("Reorder created."))
            return redirect("reorder_list")
    else:
        form = ReorderForm(initial={"quantity": suggested_quantity}, account=request.account)

    return render(
        request,
        "books/reorder_form.html",
        {
            "form": form,
            "book": book,
        },
    )


@login_required
@permission_required("books.change_reorder", raise_exception=True)
@require_POST
def reorder_update_status(request, id, action):
    reorder = get_object_or_404(Reorder, id=id, account=request.account)

    transitions = {
        "ordered": (Reorder.STATUS_PENDING, Reorder.STATUS_ORDERED),
        "received": (Reorder.STATUS_ORDERED, Reorder.STATUS_RECEIVED),
        "cancelled": (None, Reorder.STATUS_CANCELLED),
    }

    if action not in transitions:
        return redirect("reorder_list")

    required_status, new_status = transitions[action]

    if action == "cancelled":
        if reorder.status in (Reorder.STATUS_RECEIVED, Reorder.STATUS_CANCELLED):
            messages.error(request, gettext("This reorder can't be updated from its current status."))
            return redirect("reorder_list")
    elif reorder.status != required_status:
        messages.error(request, gettext("This reorder can't be updated from its current status."))
        return redirect("reorder_list")

    reorder.status = new_status

    if new_status == Reorder.STATUS_RECEIVED:
        reorder.received_at = timezone.now()
        reorder.save(update_fields=["status", "received_at"])
        book = _adjust_stock(reorder.book_id, reorder.quantity, request.user, request.account)
        messages.success(request, gettext("Reorder received and stock updated."))
        _notify_stock_level(request, book)
    else:
        reorder.save(update_fields=["status"])

        if new_status == Reorder.STATUS_ORDERED:
            messages.success(request, gettext("Reorder marked as ordered."))
        else:
            messages.success(request, gettext("Reorder cancelled."))

    _send_reorder_status_email(request.user, reorder)

    return redirect("reorder_list")


@login_required
@permission_required("books.delete_reorder", raise_exception=True)
def reorder_delete(request, id):
    reorder = get_object_or_404(Reorder, id=id, account=request.account)

    if reorder.status != Reorder.STATUS_CANCELLED:
        messages.error(request, gettext("Only cancelled reorders can be deleted."))
        return redirect("reorder_list")

    if request.method == "POST":
        reorder.delete()
        messages.success(request, gettext("Reorder deleted."))
        return redirect("reorder_list")

    return render(
        request,
        "books/confirm_delete.html",
        {
            "object_type": gettext("reorder"),
            "object_name": f"{reorder.book.title} ({reorder.quantity})",
            "cancel_url": reverse("reorder_list"),
        },
    )


@login_required
@permission_required("books.add_book", raise_exception=True)
def import_books_csv(request):
    results = None

    if request.method == "POST" and request.FILES.get("csv_file"):
        f = request.FILES["csv_file"]

        if not f.name.lower().endswith(".csv"):
            messages.error(request, gettext("Please upload a .csv file."))
            return redirect("import_books_csv")

        if f.size > settings.CSV_IMPORT_MAX_SIZE_BYTES:
            messages.error(request, gettext("File is too large (maximum 5 MB)."))
            return redirect("import_books_csv")

        try:
            text = f.read().decode("utf-8-sig")
        except UnicodeDecodeError:
            messages.error(request, gettext("File must be UTF-8 encoded."))
            return redirect("import_books_csv")

        reader = csv.DictReader(text.splitlines())

        # Normalise header names to lower-case, strip spaces
        reader.fieldnames = [h.strip().lower() for h in (reader.fieldnames or [])]

        created = updated = skipped = 0
        errors = []

        for i, row in enumerate(reader, start=2):  # row 1 is header
            title = (row.get("title") or "").strip()
            if not title:
                errors.append(gettext("Row %(n)s: title is required.") % {"n": i})
                skipped += 1
                continue

            isbn = (row.get("isbn") or "").strip() or None
            subtitle = (row.get("subtitle") or "").strip()
            publisher = (row.get("publisher") or "").strip()
            published_date_raw = (row.get("published date") or row.get("published_date") or "").strip()
            category_name = (row.get("category") or "").strip()
            expense_raw = (row.get("distribution expense") or row.get("distribution_expense") or "").strip()

            # Resolve category
            category = None
            if category_name:
                category, _ = Category.objects.get_or_create(
                    account=request.account, name=category_name,
                    defaults={"owner": request.user},
                )

            # Resolve published date
            published_date = None
            if published_date_raw:
                from datetime import date as _date
                try:
                    published_date = _date.fromisoformat(published_date_raw)
                except ValueError:
                    pass

            # Resolve distribution expense
            from decimal import InvalidOperation
            try:
                expense = Decimal(expense_raw) if expense_raw else Decimal("0")
            except InvalidOperation:
                expense = Decimal("0")

            # Find existing book by ISBN or create new
            book = None
            is_new = False
            if isbn:
                book = Book.objects.filter(account=request.account, isbn=isbn).first()

            if book is None:
                book = Book(owner=request.user, account=request.account)
                is_new = True

            if is_new and not published_date:
                errors.append(
                    gettext("Row %(n)s: a valid published date (YYYY-MM-DD) is required for new books.") % {"n": i}
                )
                skipped += 1
                continue

            if is_new and not category:
                errors.append(gettext("Row %(n)s: category is required for new books.") % {"n": i})
                skipped += 1
                continue

            book.title = title
            book.subtitle = subtitle
            book.publisher = publisher
            book.distribution_expense = expense
            if published_date:
                book.published_date = published_date
            if category:
                book.category = category
            if isbn:
                book.isbn = isbn
            book.save()

            # Resolve authors
            authors_raw = (row.get("authors") or "").strip()
            if authors_raw:
                author_objs = []
                for name in authors_raw.split(","):
                    name = name.strip()
                    if name:
                        author, _ = Author.objects.get_or_create(
                            account=request.account, name=name,
                            defaults={"owner": request.user},
                        )
                        author_objs.append(author)
                if author_objs:
                    book.authors.set(author_objs)

            if is_new:
                created += 1
            else:
                updated += 1

        results = {"created": created, "updated": updated, "skipped": skipped, "errors": errors}

    return render(request, "books/import_books.html", {"results": results})


@login_required
@permission_required("books.view_book", raise_exception=True)
def import_books_template(request):
    """Return an empty CSV with the correct headers for the user to fill in."""
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-import-template.csv"'
    writer = csv.writer(response)
    writer.writerow(["ISBN", "Title", "Subtitle", "Authors", "Publisher", "Published Date", "Category", "Distribution Expense"])
    return response


_PUBLISH_DATE_FORMATS = ["%B %d, %Y", "%b %d, %Y", "%B %Y", "%b %Y", "%Y-%m-%d", "%Y"]


def _parse_publish_date(value):
    value = (value or "").strip()
    if not value:
        return ""

    for fmt in _PUBLISH_DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue

    year_match = re.search(r"\b(1[5-9]\d{2}|20\d{2})\b", value)
    if year_match:
        return date(int(year_match.group(1)), 1, 1).isoformat()

    return ""


@login_required
@require_POST
def isbn_lookup(request):
    isbn = request.POST.get("isbn", "").strip()
    if not isbn:
        return JsonResponse({"error": gettext("ISBN is required.")}, status=400)

    cooldown_key = f"isbn_lookup_cooldown:{request.user.id}"
    if not cache.add(cooldown_key, True, timeout=settings.ISBN_LOOKUP_COOLDOWN_SECONDS):
        return JsonResponse({"error": gettext("Please wait a moment before looking up another ISBN.")}, status=429)

    try:
        data = lookup_isbn(isbn)
    except IsbnLookupError:
        return JsonResponse({"error": gettext("No book found for this ISBN.")}, status=404)

    authors_payload = []
    for name in data["authors"]:
        author, _ = Author.objects.get_or_create(
            account=request.account, name=name, defaults={"owner": request.user},
        )
        authors_payload.append({"id": author.id, "name": author.name})

    return JsonResponse({
        "title": data["title"],
        "subtitle": data["subtitle"],
        "publisher": data["publishers"][0] if data["publishers"] else "",
        "published_date": _parse_publish_date(data["publish_date"]),
        "cover_url": data["cover_url"],
        "authors": authors_payload,
    })


@login_required
@permission_required("books.view_book", raise_exception=True)
def export_books_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-books.csv"'

    writer = csv.writer(response)
    writer.writerow(_book_export_headers())

    for row in _book_export_rows(_filtered_books_for_export(request)):
        writer.writerow(row)

    return response


@login_required
@permission_required("books.view_book", raise_exception=True)
def export_books_excel(request):
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Books"
    worksheet.append(_book_export_headers())

    for row in _book_export_rows(_filtered_books_for_export(request)):
        worksheet.append(row)

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 2,
            40,
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="rumi-press-books.xlsx"'

    return response


@login_required
@permission_required("books.view_book", raise_exception=True)
def export_books_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _register_pdf_fonts()
    body_font, bold_font = _pdf_fonts()

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = bold_font

    elements = [
        Paragraph(_pdf_text(gettext("Rumi Press Books")), title_style),
        Spacer(1, 12),
    ]

    rows = [[_pdf_text(value) for value in _book_export_headers()]]

    for row in _book_export_rows(_filtered_books_for_export(request)):
        rows.append([_pdf_text(value) for value in row])

    table = Table(
        rows,
        repeatRows=1,
        colWidths=[76, 120, 105, 120, 100, 70, 90, 70],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1f1f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, -1), body_font),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements.append(table)
    document.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-books.pdf"'

    return response


@login_required
@permission_required("books.view_sale", raise_exception=True)
def export_sales_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-sales.csv"'

    writer = csv.writer(response)
    writer.writerow(_sale_export_headers())

    for row in _sale_export_rows(_sale_filters(request)):
        writer.writerow(row)

    return response


@login_required
@permission_required("books.view_sale", raise_exception=True)
def export_sales_excel(request):
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales"
    worksheet.append(_sale_export_headers())

    for row in _sale_export_rows(_sale_filters(request)):
        worksheet.append(row)

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 2,
            40,
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="rumi-press-sales.xlsx"'

    return response


@login_required
@permission_required("books.view_sale", raise_exception=True)
def export_sales_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _register_pdf_fonts()
    body_font, bold_font = _pdf_fonts()

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = bold_font

    elements = [
        Paragraph(_pdf_text(gettext("Rumi Press Sales")), title_style),
        Spacer(1, 12),
    ]

    rows = [[_pdf_text(value) for value in _sale_export_headers()]]

    for row in _sale_export_rows(_sale_filters(request)):
        rows.append([_pdf_text(value) for value in row])

    table = Table(
        rows,
        repeatRows=1,
        colWidths=[70, 160, 100, 70, 70, 70, 100],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1f1f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, -1), body_font),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements.append(table)
    document.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-sales.pdf"'

    return response


@login_required
@permission_required("books.view_reorder", raise_exception=True)
def export_reorders_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-reorders.csv"'

    writer = csv.writer(response)
    writer.writerow(_reorder_export_headers())

    for row in _reorder_export_rows(_reorder_filters(request)):
        writer.writerow(row)

    return response


@login_required
@permission_required("books.view_reorder", raise_exception=True)
def export_reorders_excel(request):
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reorders"
    worksheet.append(_reorder_export_headers())

    for row in _reorder_export_rows(_reorder_filters(request)):
        worksheet.append(row)

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 2,
            40,
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="rumi-press-reorders.xlsx"'

    return response


@login_required
@permission_required("books.view_reorder", raise_exception=True)
def export_reorders_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _register_pdf_fonts()
    body_font, bold_font = _pdf_fonts()

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = bold_font

    elements = [
        Paragraph(_pdf_text(gettext("Rumi Press Reorders")), title_style),
        Spacer(1, 12),
    ]

    rows = [[_pdf_text(value) for value in _reorder_export_headers()]]

    for row in _reorder_export_rows(_reorder_filters(request)):
        rows.append([_pdf_text(value) for value in row])

    table = Table(
        rows,
        repeatRows=1,
        colWidths=[60, 120, 90, 45, 55, 60, 50, 100, 60],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1f1f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, -1), body_font),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements.append(table)
    document.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-reorders.pdf"'

    return response


@login_required
@permission_required("books.view_invoice", raise_exception=True)
def export_invoices_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-invoices.csv"'

    writer = csv.writer(response)
    writer.writerow(_invoice_export_headers())

    for row in _invoice_export_rows(_invoice_filters(request)):
        writer.writerow(row)

    return response


@login_required
@permission_required("books.view_invoice", raise_exception=True)
def export_invoices_excel(request):
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoices"
    worksheet.append(_invoice_export_headers())

    for row in _invoice_export_rows(_invoice_filters(request)):
        worksheet.append(row)

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(
            max_length + 2,
            40,
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="rumi-press-invoices.xlsx"'

    return response


@login_required
@permission_required("books.view_invoice", raise_exception=True)
def export_invoices_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _register_pdf_fonts()
    body_font, bold_font = _pdf_fonts()

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = bold_font

    elements = [
        Paragraph(_pdf_text(gettext("Rumi Press Invoices")), title_style),
        Spacer(1, 12),
    ]

    rows = [[_pdf_text(value) for value in _invoice_export_headers()]]

    for row in _invoice_export_rows(_invoice_filters(request)):
        rows.append([_pdf_text(value) for value in row])

    table = Table(
        rows,
        repeatRows=1,
        colWidths=[70, 140, 60, 60, 60, 70, 60, 70, 60],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1f1f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, -1), body_font),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements.append(table)
    document.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-invoices.pdf"'

    return response


def _next_invoice_number(account):
    from datetime import date as _date
    year = _date.today().year
    prefix = f"INV-{year}-"
    last = (
        Invoice.objects.filter(account=account, invoice_number__startswith=prefix)
        .order_by("-invoice_number")
        .values_list("invoice_number", flat=True)
        .first()
    )
    seq = (int(last.split("-")[-1]) + 1) if last else 1
    return f"{prefix}{seq:04d}"


def _invoice_filters(request):
    invoices = Invoice.objects.filter(account=request.account).prefetch_related("items")
    status = request.GET.get("status")
    today = timezone.now().date()

    if status == "overdue":
        invoices = invoices.filter(due_date__lt=today, due_date__isnull=False).exclude(status=Invoice.STATUS_PAID)
    elif status in dict(Invoice.STATUS_CHOICES):
        invoices = invoices.filter(status=status)

    search = request.GET.get("q", "").strip()
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search)
            | Q(customer_name__icontains=search)
            | Q(customer_email__icontains=search)
        )

    start_date = request.GET.get("start_date", "").strip()
    if start_date:
        invoices = invoices.filter(invoice_date__gte=start_date)

    end_date = request.GET.get("end_date", "").strip()
    if end_date:
        invoices = invoices.filter(invoice_date__lte=end_date)

    customer_id = request.GET.get("customer", "")
    if customer_id.isdigit():
        invoices = invoices.filter(customer_id=customer_id)

    return invoices


def _active_invoice_filters(request, query_params):
    def remove(*keys):
        params = query_params.copy()
        for key in keys:
            params.pop(key, None)
        return params.urlencode()

    filters = []

    search = request.GET.get("q", "").strip()
    if search:
        filters.append({"label": gettext('Search: "%(q)s"') % {"q": search}, "url": remove("q")})

    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    if start_date or end_date:
        filters.append({
            "label": gettext("Date: %(start)s – %(end)s") % {
                "start": start_date or gettext("any"),
                "end": end_date or gettext("any"),
            },
            "url": remove("start_date", "end_date"),
        })

    customer_id = request.GET.get("customer", "")
    if customer_id.isdigit():
        customer = Customer.objects.filter(id=customer_id, account=request.account).first()
        if customer:
            filters.append({"label": gettext("Customer: %(name)s") % {"name": customer.name}, "url": remove("customer")})

    return filters


@login_required
@permission_required("books.view_invoice", raise_exception=True)
def invoice_list(request):
    invoices = _invoice_filters(request)
    status = request.GET.get("status")
    today = timezone.now().date()

    overdue_count = Invoice.objects.filter(
        account=request.account,
        due_date__lt=today,
        due_date__isnull=False,
    ).exclude(status=Invoice.STATUS_PAID).count()

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    paginator = Paginator(invoices, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    status_base_params = query_params.copy()
    status_base_params.pop("status", None)

    def status_url(value):
        params = status_base_params.copy()
        if value:
            params["status"] = value
        return "?" + params.urlencode()

    status_links = [
        {"value": "", "label": gettext("All"), "url": status_url(""), "active": not status}
    ] + [
        {"value": value, "label": label, "url": status_url(value), "active": status == value}
        for value, label in Invoice.STATUS_CHOICES
    ]
    overdue_link = {"url": status_url("overdue"), "active": status == "overdue"}

    return render(request, "books/invoice_list.html", {
        "invoices": page_obj.object_list,
        "page_obj": page_obj,
        "selected_status": status,
        "status_links": status_links,
        "overdue_link": overdue_link,
        "overdue_count": overdue_count,
        "query_string": query_string,
        "filters": request.GET,
        "active_filters": _active_invoice_filters(request, query_params),
        "result_count_text": gettext("%(count)s invoice(s) found") % {"count": paginator.count},
        "has_any_invoices": Invoice.objects.filter(account=request.account).exists(),
    })


@login_required
@permission_required("books.add_invoice", raise_exception=True)
def invoice_create(request):
    customers = Customer.objects.filter(account=request.account).values("id", "name", "email", "address")
    customers_json = _safe_json(list(customers))

    initial = {}
    customer_id = request.GET.get("customer", "")
    if customer_id.isdigit():
        customer = Customer.objects.filter(id=customer_id, account=request.account).first()
        if customer:
            initial = {
                "customer": customer.id,
                "customer_name": customer.name,
                "customer_email": customer.email,
                "customer_address": customer.address,
            }

    form = InvoiceForm(account=request.account, initial=initial)

    if request.method == "POST":
        form = InvoiceForm(request.POST, account=request.account)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.owner = request.user
            invoice.account = request.account
            invoice.invoice_number = _next_invoice_number(request.account)
            invoice.save()
            messages.success(request, gettext("Invoice %(number)s created.") % {"number": invoice.invoice_number})
            return redirect("invoice_detail", id=invoice.id)

    return render(request, "books/invoice_form.html", {"form": form, "customers_json": customers_json})


@login_required
@permission_required("books.view_invoice", raise_exception=True)
def invoice_detail(request, id):
    invoice = get_object_or_404(Invoice, id=id, account=request.account)
    item_form = InvoiceItemForm(account=request.account)

    return render(request, "books/invoice_detail.html", {
        "invoice": invoice,
        "item_form": item_form,
    })


@login_required
@permission_required("books.add_invoiceitem", raise_exception=True)
@require_POST
def invoice_item_add(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id, account=request.account)
    form = InvoiceItemForm(request.POST, account=request.account)

    if form.is_valid():
        item = form.save(commit=False)
        item.invoice = invoice
        item.save()
        messages.success(request, gettext("Item added."))
    else:
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)

    return redirect("invoice_detail", id=invoice.id)


@login_required
@permission_required("books.delete_invoiceitem", raise_exception=True)
@require_POST
def invoice_item_delete(request, id):
    item = get_object_or_404(InvoiceItem, id=id, invoice__account=request.account)
    invoice_id = item.invoice_id
    item.delete()
    messages.success(request, gettext("Item removed."))
    return redirect("invoice_detail", id=invoice_id)


def _invoice_status_transitions():
    return {
        "sent": (Invoice.STATUS_DRAFT, Invoice.STATUS_SENT),
        "paid": (Invoice.STATUS_SENT, Invoice.STATUS_PAID),
    }


def _advance_invoice_status(invoice, action, request):
    transitions = _invoice_status_transitions()
    if action not in transitions:
        return False

    required, new_status = transitions[action]
    if invoice.status != required:
        return False

    invoice.status = new_status
    invoice.save(update_fields=["status"])

    if new_status == Invoice.STATUS_SENT and invoice.customer_email:
        portal_note = ""
        if invoice.customer_id:
            portal_link = request.build_absolute_uri(reverse("customer_portal_login"))
            portal_note = (
                f"\nYou can view this invoice and your billing history anytime at:\n{portal_link}\n"
            )

        send_mail(
            subject=f"Invoice {invoice.invoice_number} from Rumi Press",
            message=(
                f"Dear {invoice.customer_name},\n\n"
                f"Please find invoice {invoice.invoice_number} attached.\n"
                f"Amount due: {invoice.grand_total} {invoice.currency}\n"
                + (f"Due date: {invoice.due_date}\n" if invoice.due_date else "")
                + portal_note
                + (f"\n{invoice.note}" if invoice.note else "")
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[invoice.customer_email],
            fail_silently=True,
        )

    return True


def _is_safe_next_url(request, next_url):
    return bool(next_url) and url_has_allowed_host_and_scheme(
        url=next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure(),
    )


@login_required
@permission_required("books.change_invoice", raise_exception=True)
@require_POST
def invoice_update_status(request, id, action):
    invoice = get_object_or_404(Invoice, id=id, account=request.account)

    if action not in _invoice_status_transitions():
        return redirect("invoice_list")

    if not _advance_invoice_status(invoice, action, request):
        messages.error(request, gettext("Cannot update invoice from its current status."))
        return redirect("invoice_detail", id=invoice.id)

    messages.success(request, gettext("Invoice marked as %(status)s.") % {"status": invoice.get_status_display()})
    next_url = request.POST.get("next", "")
    if _is_safe_next_url(request, next_url):
        return redirect(next_url)
    return redirect("invoice_detail", id=invoice.id)


@login_required
@permission_required("books.change_invoice", raise_exception=True)
@require_POST
def invoice_bulk_update(request):
    action = request.POST.get("action")
    ids = request.POST.getlist("ids")
    next_url = request.POST.get("next", "")

    if action not in _invoice_status_transitions() or not ids:
        messages.error(request, gettext("Select at least one invoice and an action."))
    else:
        updated = 0
        for invoice in Invoice.objects.filter(account=request.account, id__in=ids):
            if _advance_invoice_status(invoice, action, request):
                updated += 1
        skipped = len(ids) - updated

        if updated:
            messages.success(request, gettext("%(count)d invoice(s) updated.") % {"count": updated})
        if skipped:
            messages.warning(
                request,
                gettext("%(count)d invoice(s) skipped (incompatible status).") % {"count": skipped},
            )

    if _is_safe_next_url(request, next_url):
        return redirect(next_url)
    return redirect("invoice_list")


@login_required
@permission_required("books.delete_invoice", raise_exception=True)
def invoice_delete(request, id):
    invoice = get_object_or_404(Invoice, id=id, account=request.account)

    if request.method == "POST":
        invoice.delete()
        messages.success(request, gettext("Invoice deleted."))
        return redirect("invoice_list")

    return render(request, "books/confirm_delete.html", {
        "object_type": gettext("invoice"),
        "object_name": f"{invoice.invoice_number} – {invoice.customer_name}",
        "cancel_url": reverse("invoice_detail", args=[invoice.id]),
    })


INVOICE_AGING_BUCKETS = [
    ("current", gettext_lazy("Current (not yet due)")),
    ("0_30", gettext_lazy("1–30 days overdue")),
    ("31_60", gettext_lazy("31–60 days overdue")),
    ("60_plus", gettext_lazy("60+ days overdue")),
]


def _invoice_aging_data(account):
    today = timezone.now().date()
    invoices = (
        Invoice.objects.filter(account=account)
        .exclude(status=Invoice.STATUS_PAID)
        .prefetch_related("items")
        .order_by("due_date")
    )

    buckets = {key: {"label": label, "invoices": [], "total": Decimal(0)} for key, label in INVOICE_AGING_BUCKETS}

    for invoice in invoices:
        total = invoice.grand_total

        if invoice.due_date is None or invoice.due_date >= today:
            key, days_overdue = "current", 0
        else:
            days_overdue = (today - invoice.due_date).days
            if days_overdue <= 30:
                key = "0_30"
            elif days_overdue <= 60:
                key = "31_60"
            else:
                key = "60_plus"

        buckets[key]["invoices"].append({"invoice": invoice, "days_overdue": days_overdue, "total": total})
        buckets[key]["total"] += total

    grand_total = sum((bucket["total"] for bucket in buckets.values()), Decimal(0))
    return buckets, grand_total


@login_required
@permission_required("books.view_invoice", raise_exception=True)
def invoice_aging_report(request):
    buckets, grand_total = _invoice_aging_data(request.account)

    return render(request, "books/invoice_aging_report.html", {
        "buckets": buckets,
        "grand_total": grand_total,
    })


def _invoice_pdf_response(invoice):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _register_pdf_fonts()
    body_font, bold_font = _pdf_fonts()

    items = invoice.items.all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=48, rightMargin=48, topMargin=48, bottomMargin=48)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = bold_font
    styles["Normal"].fontName = body_font

    elements = [
        Paragraph(_pdf_text(f"INVOICE – {invoice.invoice_number}"), styles["Title"]),
        Spacer(1, 8),
        Paragraph(_pdf_text(f"{gettext('Customer')}: {invoice.customer_name}"), styles["Normal"]),
    ]

    if invoice.customer_address:
        elements.append(Paragraph(_pdf_text(invoice.customer_address.replace("\n", " | ")), styles["Normal"]))
    if invoice.customer_email:
        elements.append(Paragraph(_pdf_text(f"{gettext('Email')}: {invoice.customer_email}"), styles["Normal"]))

    elements += [
        Paragraph(_pdf_text(f"{gettext('Date')}: {invoice.invoice_date}"), styles["Normal"]),
        Spacer(1, 12),
    ]

    if invoice.due_date:
        elements.append(Paragraph(_pdf_text(f"{gettext('Due')}: {invoice.due_date}"), styles["Normal"]))

    rows = [[
        _pdf_text(gettext("Description")),
        _pdf_text(gettext("Qty")),
        _pdf_text(gettext("Unit Price")),
        _pdf_text(gettext("Tax %")),
        _pdf_text(gettext("Total")),
    ]]
    for item in items:
        rows.append([
            _pdf_text(item.description),
            str(item.quantity),
            str(item.unit_price),
            str(item.tax_rate),
            str(item.total),
        ])

    rows.append(["", "", "", _pdf_text(gettext("Grand Total")), _pdf_text(f"{invoice.grand_total} {invoice.currency}")])

    table = Table(rows, repeatRows=1, colWidths=[220, 50, 80, 60, 80])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1f1f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), body_font),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -2), 0.25, colors.grey),
        ("FONTNAME", (-2, -1), (-1, -1), bold_font),
        ("LINEABOVE", (-2, -1), (-1, -1), 1, colors.black),
    ]))

    elements += [Spacer(1, 16), table]

    if invoice.note:
        elements += [Spacer(1, 12), Paragraph(_pdf_text(f"{gettext('Note')}: {invoice.note}"), styles["Normal"])]

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice-{invoice.invoice_number}.pdf"'
    return response


@login_required
@permission_required("books.view_invoice", raise_exception=True)
def invoice_pdf(request, id):
    invoice = get_object_or_404(Invoice, id=id, account=request.account)
    return _invoice_pdf_response(invoice)


@login_required
@permission_required("books.view_printrun", raise_exception=True)
def print_run_list(request):
    runs = PrintRun.objects.filter(account=request.account).select_related("book", "book__category")
    paginator = Paginator(runs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "books/print_run_list.html", {
        "runs": page_obj.object_list,
        "page_obj": page_obj,
    })


@login_required
@permission_required("books.add_printrun", raise_exception=True)
def print_run_create(request, book_id):
    book = get_object_or_404(Book, id=book_id, account=request.account)

    if request.method == "POST":
        form = PrintRunForm(request.POST)
        if form.is_valid():
            run = form.save(commit=False)
            run.owner = request.user
            run.account = request.account
            run.book = book
            run.save()
            messages.success(request, gettext("Print run recorded."))
            return redirect("print_run_list")
    else:
        form = PrintRunForm()

    return render(request, "books/print_run_form.html", {"form": form, "book": book})


@login_required
@permission_required("books.change_printrun", raise_exception=True)
@require_POST
def print_run_complete(request, id):
    run = get_object_or_404(PrintRun, id=id, account=request.account)

    if run.status != PrintRun.STATUS_PENDING:
        messages.error(request, gettext("This print run is already completed."))
        return redirect("print_run_list")

    run.status = PrintRun.STATUS_COMPLETED
    run.completed_at = timezone.now()
    run.save(update_fields=["status", "completed_at"])

    book = _adjust_stock(run.book_id, run.quantity, request.user, request.account)
    StockAdjustment.objects.create(
        owner=request.user,
        account=request.account,
        book=run.book,
        change=run.quantity,
        resulting_stock=book.stock_on_hand,
        reason=StockAdjustment.REASON_PRODUCTION,
        note=gettext("Print run completed (%(date)s)") % {"date": run.run_date},
    )

    messages.success(request, gettext("Print run marked complete; stock updated."))
    _notify_stock_level(request, book)
    return redirect("print_run_list")


@login_required
@permission_required("books.delete_printrun", raise_exception=True)
def print_run_delete(request, id):
    run = get_object_or_404(PrintRun, id=id, account=request.account)

    if run.status == PrintRun.STATUS_COMPLETED:
        messages.error(request, gettext("Completed print runs can't be deleted; the stock adjustment must be reversed manually."))
        return redirect("print_run_list")

    if request.method == "POST":
        run.delete()
        messages.success(request, gettext("Print run deleted."))
        return redirect("print_run_list")

    return render(request, "books/confirm_delete.html", {
        "object_type": gettext("print run"),
        "object_name": f"{run.book.title} – {run.run_date}",
        "cancel_url": reverse("print_run_list"),
    })


@login_required
@permission_required("books.view_royaltyrate", raise_exception=True)
def royalty_list(request):
    rates = RoyaltyRate.objects.filter(account=request.account).select_related("book", "author")
    return render(request, "books/royalty_list.html", {"rates": rates})


@login_required
@permission_required("books.add_royaltyrate", raise_exception=True)
def royalty_create(request):
    form = RoyaltyRateForm(account=request.account)

    if request.method == "POST":
        form = RoyaltyRateForm(request.POST, account=request.account)
        if form.is_valid():
            rate = form.save(commit=False)
            rate.owner = request.user
            rate.account = request.account
            rate.save()
            messages.success(request, gettext("Royalty rate added."))
            return redirect("royalty_list")

    return render(request, "books/royalty_form.html", {"form": form})


@login_required
@permission_required("books.delete_royaltyrate", raise_exception=True)
def royalty_delete(request, id):
    rate = get_object_or_404(RoyaltyRate, id=id, account=request.account)

    if request.method == "POST":
        rate.delete()
        messages.success(request, gettext("Royalty rate deleted."))
        return redirect("royalty_list")

    return render(request, "books/confirm_delete.html", {
        "object_type": gettext("royalty rate"),
        "object_name": str(rate),
        "cancel_url": reverse("royalty_list"),
    })


@login_required
@permission_required("books.view_royaltyrate", raise_exception=True)
def royalty_report(request):
    rates = RoyaltyRate.objects.filter(account=request.account).select_related("book", "author")

    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()

    sales_qs = Sale.objects.filter(account=request.account)
    if start_date:
        sales_qs = sales_qs.filter(sale_date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(sale_date__lte=end_date)

    revenue_by_book = {
        item["book_id"]: item["revenue"] or 0
        for item in sales_qs.values("book_id").annotate(revenue=Sum(REVENUE_EXPRESSION))
    }

    rows = []
    for rate in rates:
        total_revenue = Decimal(str(revenue_by_book.get(rate.book_id, 0)))
        royalty_amount = total_revenue * rate.rate / 100
        rows.append({
            "book": rate.book,
            "author": rate.author,
            "rate": rate.rate,
            "effective_from": rate.effective_from,
            "total_revenue": total_revenue,
            "royalty_amount": royalty_amount,
        })

    payment_totals = {}
    for item in (
        RoyaltyPayment.objects.filter(account=request.account)
        .values("author__name", "currency")
        .annotate(total_paid=Sum("amount"))
        .order_by("author__name")
    ):
        payment_totals.setdefault(item["author__name"], []).append({
            "currency": item["currency"],
            "total_paid": item["total_paid"],
        })
    payment_rows = [{"author": name, "totals": totals} for name, totals in payment_totals.items()]

    return render(request, "books/royalty_report.html", {
        "rows": rows,
        "start_date": start_date,
        "end_date": end_date,
        "payment_rows": payment_rows,
    })


@login_required
@permission_required("books.view_royaltypayment", raise_exception=True)
def royalty_payment_list(request):
    payments = RoyaltyPayment.objects.filter(account=request.account).select_related("author")
    return render(request, "books/royalty_payment_list.html", {"payments": payments})


@login_required
@permission_required("books.add_royaltypayment", raise_exception=True)
def royalty_payment_create(request):
    form = RoyaltyPaymentForm(account=request.account)

    if request.method == "POST":
        form = RoyaltyPaymentForm(request.POST, account=request.account)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.owner = request.user
            payment.account = request.account
            payment.save()
            messages.success(request, gettext("Royalty payment recorded."))
            return redirect("royalty_payment_list")

    return render(request, "books/royalty_payment_form.html", {"form": form})


@login_required
@permission_required("books.delete_royaltypayment", raise_exception=True)
def royalty_payment_delete(request, id):
    payment = get_object_or_404(RoyaltyPayment, id=id, account=request.account)

    if request.method == "POST":
        payment.delete()
        messages.success(request, gettext("Royalty payment deleted."))
        return redirect("royalty_payment_list")

    return render(request, "books/confirm_delete.html", {
        "object_type": gettext("royalty payment"),
        "object_name": str(payment),
        "cancel_url": reverse("royalty_payment_list"),
    })


# ---------------------------------------------------------------------------
# Location views
# ---------------------------------------------------------------------------

@login_required
@permission_required("books.view_location", raise_exception=True)
def location_list(request):
    locations = Location.objects.filter(account=request.account).annotate(
        total_stock=Sum("stock_levels__quantity"),
    )
    return render(request, "books/location_list.html", {"locations": locations})


@login_required
@permission_required("books.add_location", raise_exception=True)
def location_create(request):
    if request.method == "POST":
        form = LocationForm(request.POST)
        if form.is_valid():
            loc = form.save(commit=False)
            loc.owner = request.user
            loc.account = request.account
            if loc.is_default:
                Location.objects.filter(account=request.account, is_default=True).update(is_default=False)
            loc.save()
            messages.success(request, gettext("Location created."))
            return redirect("location_list")
    else:
        form = LocationForm()
    return render(request, "books/location_form.html", {"form": form, "title": gettext("Add location")})


@login_required
@permission_required("books.change_location", raise_exception=True)
def location_update(request, id):
    loc = get_object_or_404(Location, id=id, account=request.account)
    if request.method == "POST":
        form = LocationForm(request.POST, instance=loc)
        if form.is_valid():
            updated = form.save(commit=False)
            if updated.is_default:
                Location.objects.filter(account=request.account, is_default=True).exclude(id=id).update(is_default=False)
            updated.save()
            messages.success(request, gettext("Location updated."))
            return redirect("location_list")
    else:
        form = LocationForm(instance=loc)
    return render(request, "books/location_form.html", {"form": form, "title": gettext("Edit location"), "object": loc})


@login_required
@permission_required("books.delete_location", raise_exception=True)
def location_delete(request, id):
    loc = get_object_or_404(Location, id=id, account=request.account)
    if request.method == "POST":
        loc.delete()
        messages.success(request, gettext("Location deleted."))
        return redirect("location_list")
    return render(request, "books/confirm_delete.html", {
        "object_type": gettext("location"),
        "object_name": loc.name,
        "cancel_url": reverse("location_list"),
    })


@login_required
@permission_required("books.view_stocklevel", raise_exception=True)
def location_stock(request, id):
    loc = get_object_or_404(Location, id=id, account=request.account)
    levels = StockLevel.objects.filter(location=loc, account=request.account).select_related("book")
    return render(request, "books/location_stock.html", {"location": loc, "levels": levels})


@login_required
@permission_required("books.change_stocklevel", raise_exception=True)
def stock_transfer(request):
    form = StockTransferForm(request.POST or None, account=request.account)
    if request.method == "POST" and form.is_valid():
        book = form.cleaned_data["book"]
        from_loc = form.cleaned_data["from_location"]
        to_loc = form.cleaned_data["to_location"]
        qty = form.cleaned_data["quantity"]

        if from_loc == to_loc:
            form.add_error("to_location", gettext("Source and destination must differ."))
        else:
            # transaction.atomic() + select_for_update() closes the race
            # between checking "available" and writing the deduction - two
            # concurrent transfers from the same location could otherwise
            # both read the same stale quantity and the slower save() would
            # silently overwrite the faster one's update.
            with transaction.atomic():
                from_level = (
                    StockLevel.objects.select_for_update()
                    .filter(account=request.account, book=book, location=from_loc)
                    .first()
                )
                available = from_level.quantity if from_level else 0

                if available < qty:
                    form.add_error("quantity", gettext("Not enough stock at source location (available: %(n)s).") % {"n": available})
                else:
                    from_level.quantity -= qty
                    from_level.save(update_fields=["quantity"])

                    to_level, _ = StockLevel.objects.select_for_update().get_or_create(
                        account=request.account, book=book, location=to_loc,
                        defaults={"quantity": 0, "owner": request.user},
                    )
                    to_level.quantity += qty
                    to_level.save(update_fields=["quantity"])

                    messages.success(request, gettext("Transferred %(qty)s copies of «%(book)s» from %(from)s to %(to)s.") % {
                        "qty": qty, "book": book.title, "from": from_loc.name, "to": to_loc.name,
                    })
                    return redirect("location_list")

    return render(request, "books/stock_transfer_form.html", {"form": form})


# ---------------------------------------------------------------------------
# Integration views
# ---------------------------------------------------------------------------

@login_required
@permission_required("books.view_integration", raise_exception=True)
def integration_list(request):
    integrations = Integration.objects.filter(account=request.account)
    return render(request, "books/integration_list.html", {"integrations": integrations})


@login_required
@permission_required("books.add_integration", raise_exception=True)
def integration_create(request):
    if request.method == "POST":
        form = IntegrationForm(request.POST)
        if form.is_valid():
            intg = form.save(commit=False)
            intg.owner = request.user
            intg.account = request.account
            intg.save()
            messages.success(request, gettext("Integration saved."))
            return redirect("integration_list")
    else:
        form = IntegrationForm()
    return render(request, "books/integration_form.html", {"form": form, "title": gettext("Add integration")})


@login_required
@permission_required("books.change_integration", raise_exception=True)
def integration_update(request, id):
    intg = get_object_or_404(Integration, id=id, account=request.account)
    if request.method == "POST":
        form = IntegrationForm(request.POST, instance=intg)
        if form.is_valid():
            form.save()
            messages.success(request, gettext("Integration updated."))
            return redirect("integration_list")
    else:
        form = IntegrationForm(instance=intg)
    return render(request, "books/integration_form.html", {"form": form, "title": gettext("Edit integration"), "object": intg})


@login_required
@permission_required("books.delete_integration", raise_exception=True)
def integration_delete(request, id):
    intg = get_object_or_404(Integration, id=id, account=request.account)
    if request.method == "POST":
        intg.delete()
        messages.success(request, gettext("Integration deleted."))
        return redirect("integration_list")
    return render(request, "books/confirm_delete.html", {
        "object_type": gettext("integration"),
        "object_name": intg.name,
        "cancel_url": reverse("integration_list"),
    })


def _process_shopify_order(integration, payload):
    """Deduct stock for each line item in a Shopify order.

    Records the order ID first, inside the same transaction, so a retried
    webhook delivery for an order we've already synced is a no-op instead
    of decrementing stock twice.
    """
    order_id = str(payload.get("id") or "")

    with transaction.atomic():
        if order_id:
            _, created = ProcessedShopifyOrder.objects.get_or_create(
                integration=integration, order_id=order_id,
            )
            if not created:
                return 0

        owner = integration.owner
        account = integration.account
        line_items = payload.get("line_items", [])
        synced = 0

        for item in line_items:
            sku = item.get("sku", "").strip()
            qty = item.get("quantity", 0)
            if not sku or not qty:
                continue

            book = Book.objects.filter(account=account, isbn=sku).first()
            if book:
                _adjust_stock(book.id, -qty, owner, account)
                synced += 1

        return synced


@csrf_exempt
def shopify_webhook(request, integration_id):
    """Verify HMAC and process an incoming Shopify orders/create webhook."""
    if request.method != "POST":
        return HttpResponse(status=405)

    try:
        intg = Integration.objects.get(
            id=integration_id,
            platform=Integration.PLATFORM_SHOPIFY,
            is_active=True,
        )
    except Integration.DoesNotExist:
        return HttpResponse(status=404)

    # Verify HMAC signature
    secret = intg.webhook_secret.encode()
    body = request.body
    digest = base64.b64encode(
        _hmac.new(secret, body, hashlib.sha256).digest()
    ).decode()
    shopify_hmac = request.headers.get("X-Shopify-Hmac-Sha256", "")

    if not _hmac.compare_digest(digest, shopify_hmac):
        return HttpResponse(status=401)

    try:
        payload = json.loads(body)
    except json.JSONDecodeError:
        return HttpResponse(status=400)

    synced = _process_shopify_order(intg, payload)

    intg.orders_synced = (intg.orders_synced or 0) + (1 if synced > 0 else 0)
    intg.last_synced_at = timezone.now()
    intg.save(update_fields=["orders_synced", "last_synced_at"])

    return HttpResponse(status=200)


def _stripe_obj_get(obj, key, default=None):
    """stripe's StripeObject only supports bracket access, not dict.get()."""
    try:
        return obj[key]
    except KeyError:
        return default


@csrf_exempt
def stripe_webhook(request, integration_id):
    """Verify the signature and mark an invoice paid on checkout.session.completed."""
    if request.method != "POST":
        return HttpResponse(status=405)

    try:
        intg = Integration.objects.get(
            id=integration_id,
            platform=Integration.PLATFORM_STRIPE,
            is_active=True,
        )
    except Integration.DoesNotExist:
        return HttpResponse(status=404)

    sig_header = request.headers.get("Stripe-Signature", "")

    try:
        event = stripe.Webhook.construct_event(request.body, sig_header, intg.webhook_secret)
    except (ValueError, stripe.SignatureVerificationError):
        return HttpResponse(status=401)

    if event["type"] == "checkout.session.completed":
        session = event["data"]["object"]
        metadata = _stripe_obj_get(session, "metadata", {})
        invoice_id = _stripe_obj_get(metadata, "invoice_id")

        invoice = Invoice.objects.filter(id=invoice_id, account=intg.account).first()
        if (
            invoice is not None
            and invoice.status != Invoice.STATUS_PAID
            and _stripe_obj_get(session, "currency", "").upper() == invoice.currency
            and _stripe_obj_get(session, "amount_total") == int((invoice.grand_total * 100).to_integral_value())
        ):
            invoice.status = Invoice.STATUS_PAID
            invoice.stripe_payment_intent_id = _stripe_obj_get(session, "payment_intent", "")
            invoice.save(update_fields=["status", "stripe_payment_intent_id"])

    return HttpResponse(status=200)


# ---------------------------------------------------------------------------
# PWA views
# ---------------------------------------------------------------------------

def manifest_json(request):
    data = {
        "name": "RumiPress",
        "short_name": "RumiPress",
        "description": "Book inventory management",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#1a1a2e",
        "theme_color": "#0d6efd",
        "icons": [
            {"src": "/static/books/img/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/books/img/icon-512.png", "sizes": "512x512", "type": "image/png"},
        ],
    }
    return JsonResponse(data)


def service_worker(request):
    js = r"""
const CACHE = "rumipress-v1";
const OFFLINE_URL = "/";

self.addEventListener("install", event => {
  event.waitUntil(
    caches.open(CACHE).then(cache => cache.addAll([
      "/",
      "/books/",
      "/static/books/css/style.css",
    ]))
  );
  self.skipWaiting();
});

self.addEventListener("activate", event => {
  event.waitUntil(
    caches.keys().then(keys =>
      Promise.all(keys.filter(k => k !== CACHE).map(k => caches.delete(k)))
    )
  );
  self.clients.claim();
});

self.addEventListener("fetch", event => {
  if (event.request.method !== "GET") return;
  event.respondWith(
    fetch(event.request)
      .then(response => {
        const clone = response.clone();
        caches.open(CACHE).then(cache => cache.put(event.request, clone));
        return response;
      })
      .catch(() => caches.match(event.request).then(r => r || caches.match(OFFLINE_URL)))
  );
});
""".strip()
    return HttpResponse(js, content_type="application/javascript")


def _generate_verification_code():
    return f"{secrets.randbelow(1000000):06d}"


def _send_verification_email(user, code):
    send_mail(
        subject="Your RumiPress verification code",
        message=(
            f"Hi {user.username},\n\n"
            f"Your RumiPress email verification code is: {code}\n"
            f"This code expires in {settings.VERIFICATION_CODE_TTL_MINUTES} minutes.\n\n"
            "If you didn't request this, you can ignore this email."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=True,
    )


def _notify_owners_of_pending_activation(user):
    owner_emails = list(
        User.objects.filter(is_superuser=True, email__gt="")
        .exclude(email="")
        .values_list("email", flat=True)
    )

    if not owner_emails:
        return

    send_mail(
        subject="RumiPress: new user awaiting access code",
        message=(
            f"User '{user.username}' ({user.email}) has verified their email "
            "and is waiting for an access code to activate their account.\n\n"
            "Generate an access code in the admin site under Books > Access codes."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=owner_emails,
        fail_silently=True,
    )


def _get_pending_user(request):
    user_id = request.session.get("pending_user_id")
    if not user_id:
        return None

    try:
        return User.objects.get(id=user_id, is_active=False)
    except User.DoesNotExist:
        return None


def signup(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    form = SignupForm()

    if request.method == "POST":
        form = SignupForm(request.POST)

        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                email=form.cleaned_data["email"],
                password=form.cleaned_data["password1"],
                is_active=False,
            )

            code = _generate_verification_code()
            profile, _ = Profile.objects.get_or_create(user=user)
            profile.verification_code = code
            profile.verification_code_expires_at = timezone.now() + timedelta(
                minutes=settings.VERIFICATION_CODE_TTL_MINUTES
            )
            profile.save()

            _send_verification_email(user, code)

            request.session["pending_user_id"] = user.id
            return redirect("verify_email")

    return render(request, "registration/signup.html", {"form": form})


def verify_email(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    user = _get_pending_user(request)
    if user is None:
        return redirect("signup")

    profile, _ = Profile.objects.get_or_create(user=user)

    if profile.email_verified:
        return redirect("redeem_access_code")

    form = VerifyEmailForm()
    attempts_key = f"verify_email_attempts:{user.id}"

    if request.method == "POST":
        if request.POST.get("action") == "resend":
            code = _generate_verification_code()
            profile.verification_code = code
            profile.verification_code_expires_at = timezone.now() + timedelta(
                minutes=settings.VERIFICATION_CODE_TTL_MINUTES
            )
            profile.save()
            _send_verification_email(user, code)
            cache.delete(attempts_key)
            messages.success(request, gettext("A new verification code has been sent to your email."))
        else:
            form = VerifyEmailForm(request.POST)

            if form.is_valid():
                if cache.get(attempts_key, 0) >= settings.VERIFICATION_CODE_MAX_ATTEMPTS:
                    form.add_error(
                        "code", gettext("Too many incorrect attempts. Please request a new code."),
                    )
                else:
                    code = form.cleaned_data["code"]
                    expires_at = profile.verification_code_expires_at

                    if (
                        profile.verification_code
                        and code == profile.verification_code
                        and expires_at
                        and timezone.now() <= expires_at
                    ):
                        profile.email_verified = True
                        profile.verification_code = ""
                        profile.verification_code_expires_at = None
                        profile.save()
                        cache.delete(attempts_key)
                        _notify_owners_of_pending_activation(user)
                        return redirect("redeem_access_code")

                    cache.set(
                        attempts_key,
                        cache.get(attempts_key, 0) + 1,
                        timeout=settings.VERIFICATION_CODE_TTL_MINUTES * 60,
                    )
                    form.add_error("code", gettext("That code is invalid or has expired."))

    return render(
        request,
        "registration/verify_email.html",
        {"form": form, "email": user.email},
    )


def redeem_access_code(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    user = _get_pending_user(request)
    if user is None:
        return redirect("signup")

    profile, _ = Profile.objects.get_or_create(user=user)

    if not profile.email_verified:
        return redirect("verify_email")

    form = RedeemAccessCodeForm()
    attempts_key = f"redeem_access_code_attempts:{user.id}"

    if request.method == "POST":
        form = RedeemAccessCodeForm(request.POST)

        if form.is_valid() and cache.get(attempts_key, 0) >= settings.ACCESS_CODE_MAX_ATTEMPTS:
            form.add_error("code", gettext("Too many incorrect attempts. Please try again later."))
        elif form.is_valid():
            code_value = form.cleaned_data["code"].strip()

            try:
                access_code = AccessCode.objects.get(code__iexact=code_value)
            except AccessCode.DoesNotExist:
                access_code = None

            if access_code is None or not access_code.is_valid:
                cache.set(attempts_key, cache.get(attempts_key, 0) + 1, timeout=3600)
                form.add_error("code", gettext("That access code is invalid, used, or expired."))
            else:
                access_code.is_used = True
                access_code.used_by = user
                access_code.used_at = timezone.now()
                access_code.save()

                profile.access_code_redeemed = True
                profile.save()

                user.is_active = True
                user.save()

                admin_group = ensure_roles()["Admin"]
                user.groups.add(admin_group)

                account = Account.objects.create(name=user.username)
                AccountMembership.objects.create(account=account, user=user, role=AccountMembership.ROLE_ADMIN)

                Category.objects.get_or_create(owner=user, account=account, name="General")
                Subscription.objects.get_or_create(account=account, defaults={"user": user})

                cache.delete(attempts_key)
                del request.session["pending_user_id"]

                login(request, user, backend="django.contrib.auth.backends.ModelBackend")
                messages.success(request, gettext("Your account is active. Welcome to RumiPress!"))
                return redirect("billing_start")

    return render(request, "registration/redeem_access_code.html", {"form": form})


# ---------------------------------------------------------------------------
# Team management (invite Staff/Viewer users into the current Account)
# ---------------------------------------------------------------------------

INVITE_TTL_DAYS = 7


def _require_account_admin(request):
    membership = AccountMembership.objects.filter(
        account=request.account, user=request.user,
    ).first()
    if membership is None or membership.role != AccountMembership.ROLE_ADMIN:
        raise PermissionDenied


def _send_invite_email(invitation, account_name, accept_url):
    send_mail(
        subject="You've been invited to RumiPress",
        message=(
            f"You've been invited to join {account_name or 'a RumiPress account'} "
            f"as {invitation.role}.\n\n"
            f"Click the link below to accept:\n{accept_url}\n\n"
            f"This invite expires in {INVITE_TTL_DAYS} days.\n\n"
            "If you weren't expecting this, you can ignore this email."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[invitation.email],
        fail_silently=True,
    )


@login_required
def team_members(request):
    _require_account_admin(request)

    if request.method == "POST":
        form = InviteUserForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data["email"]
            role = form.cleaned_data["role"]

            invitation = AccountInvitation.objects.create(
                account=request.account,
                email=email,
                role=role,
                token=secrets.token_urlsafe(32),
                invited_by=request.user,
                expires_at=timezone.now() + timedelta(days=INVITE_TTL_DAYS),
            )
            accept_url = request.build_absolute_uri(
                reverse("team_accept_invite", args=[invitation.token])
            )
            _send_invite_email(invitation, request.account.name, accept_url)

            messages.success(request, gettext("Invitation sent to %(email)s.") % {"email": email})
            return redirect("team_members")
    else:
        form = InviteUserForm()

    memberships = (
        AccountMembership.objects.filter(account=request.account)
        .select_related("user")
        .order_by("user__username")
    )
    pending_invitations = AccountInvitation.objects.filter(
        account=request.account, accepted_at__isnull=True,
    ).order_by("-created_at")

    return render(request, "books/team_members.html", {
        "form": form,
        "memberships": memberships,
        "pending_invitations": pending_invitations,
    })


@login_required
@require_POST
def team_member_update_role(request, id):
    _require_account_admin(request)

    membership = get_object_or_404(AccountMembership, id=id, account=request.account)
    role = request.POST.get("role")

    if role not in dict(AccountMembership.ROLE_CHOICES):
        messages.error(request, gettext("Invalid role."))
        return redirect("team_members")

    if (
        membership.role == AccountMembership.ROLE_ADMIN
        and role != AccountMembership.ROLE_ADMIN
        and not AccountMembership.objects.filter(
            account=request.account, role=AccountMembership.ROLE_ADMIN,
        ).exclude(id=membership.id).exists()
    ):
        messages.error(request, gettext("An account must keep at least one Admin."))
        return redirect("team_members")

    membership.role = role
    membership.save(update_fields=["role"])
    sync_user_groups_for_role(membership.user, role)

    messages.success(request, gettext("Role updated."))
    return redirect("team_members")


@login_required
@require_POST
def team_member_remove(request, id):
    _require_account_admin(request)

    membership = get_object_or_404(AccountMembership, id=id, account=request.account)

    if (
        membership.role == AccountMembership.ROLE_ADMIN
        and not AccountMembership.objects.filter(
            account=request.account, role=AccountMembership.ROLE_ADMIN,
        ).exclude(id=membership.id).exists()
    ):
        messages.error(request, gettext("An account must keep at least one Admin."))
        return redirect("team_members")

    # Clear the removed user's Django Group membership too - otherwise they
    # keep acting on their old role's permissions indefinitely, since
    # nothing else revokes them once the AccountMembership row is gone.
    removed_user = membership.user
    membership.delete()
    sync_user_groups_for_role(removed_user, None)

    messages.success(request, gettext("Member removed."))
    return redirect("team_members")


@login_required
@require_POST
def team_invite_cancel(request, id):
    _require_account_admin(request)

    invitation = get_object_or_404(
        AccountInvitation, id=id, account=request.account, accepted_at__isnull=True,
    )
    invitation.delete()
    messages.success(request, gettext("Invitation cancelled."))
    return redirect("team_members")


def team_accept_invite(request, token):
    if request.user.is_authenticated:
        return redirect("dashboard")

    invitation = get_object_or_404(AccountInvitation, token=token)

    if invitation.accepted_at is not None or invitation.expires_at < timezone.now():
        return render(request, "registration/team_invite_invalid.html", status=410)

    if User.objects.filter(email__iexact=invitation.email).exists():
        return render(request, "registration/team_invite_invalid.html", {
            "already_has_account": True,
        }, status=409)

    form = AcceptInviteForm()

    if request.method == "POST":
        form = AcceptInviteForm(request.POST)

        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                email=invitation.email,
                password=form.cleaned_data["password1"],
            )
            Profile.objects.get_or_create(user=user, defaults={"email_verified": True})
            AccountMembership.objects.create(
                account=invitation.account, user=user, role=invitation.role,
            )
            sync_user_groups_for_role(user, invitation.role)

            invitation.accepted_at = timezone.now()
            invitation.save(update_fields=["accepted_at"])

            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            messages.success(request, gettext("Welcome! Your account is ready."))
            return redirect("dashboard")

    return render(request, "registration/team_accept_invite.html", {
        "form": form, "invitation": invitation,
    })


# ---------------------------------------------------------------------------
# Platform billing (iyzico subscriptions for RumiPress accounts themselves -
# distinct from the per-owner Stripe Integration used for invoice payments)
# ---------------------------------------------------------------------------

def _get_or_create_subscription(account, user):
    subscription, _ = Subscription.objects.get_or_create(account=account, defaults={"user": user})
    return subscription


@login_required
def billing_start(request):
    subscription = _get_or_create_subscription(request.account, request.user)
    if subscription.is_in_good_standing:
        return redirect("dashboard")

    form = IyzicoCustomerForm(initial={"email": request.user.email})

    if request.method == "POST":
        form = IyzicoCustomerForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            customer = {
                "name": data["name"],
                "surname": data["surname"],
                "email": data["email"],
                "gsmNumber": data["gsm_number"],
                "identityNumber": data["identity_number"],
                "billingAddress": {
                    "address": data["address"],
                    "contactName": f"{data['name']} {data['surname']}",
                    "city": data["city"],
                    "country": data["country"],
                    "zipCode": data.get("zip_code", ""),
                },
            }
            callback_url = request.build_absolute_uri(reverse("billing_callback"))

            try:
                result = iyzico_client.initialize_subscription_checkout_form(
                    settings.IYZICO_PRICING_PLAN_REFERENCE_CODE, callback_url, customer,
                )
            except iyzico_client.IyzicoError:
                messages.error(request, gettext("Couldn't start your subscription. Please try again or contact us."))
                return redirect("billing_required")

            subscription.checkout_token = result["data"]["token"]
            subscription.save(update_fields=["checkout_token"])

            return render(request, "registration/billing_checkout.html", {
                "checkout_form_content": result["data"]["checkoutFormContent"],
                "heading": gettext("Start your subscription"),
            })

    return render(request, "registration/billing_start.html", {"form": form})


@login_required
def billing_callback(request):
    subscription = _get_or_create_subscription(request.account, request.user)
    token = request.GET.get("token", "")

    try:
        result = iyzico_client.retrieve_checkout_form(token)
    except iyzico_client.IyzicoError:
        messages.error(request, gettext("Couldn't confirm your subscription. Please try again or contact us."))
        return redirect("billing_required")

    data = result.get("data", {})
    subscription.external_customer_id = data.get("customerReferenceCode", "")
    subscription.external_subscription_id = data.get("referenceCode", "")
    status = data.get("subscriptionStatus", "")
    if status in dict(Subscription.STATUS_CHOICES):
        subscription.status = status
    subscription.checkout_token = ""
    subscription.save(update_fields=[
        "external_customer_id", "external_subscription_id", "status", "checkout_token",
    ])

    if subscription.is_in_good_standing:
        messages.success(request, gettext("Your subscription is active. Welcome to RumiPress!"))
        return redirect("dashboard")

    messages.error(request, gettext("Your subscription couldn't be activated. Please try again or contact us."))
    return redirect("billing_required")


@login_required
def billing_required(request):
    subscription = _get_or_create_subscription(request.account, request.user)
    if subscription.is_in_good_standing:
        return redirect("dashboard")

    return render(request, "registration/billing_required.html", {"subscription": subscription})


@login_required
def billing_portal(request):
    subscription = _get_or_create_subscription(request.account, request.user)
    if not subscription.external_customer_id:
        return redirect("billing_start")

    callback_url = request.build_absolute_uri(reverse("billing_card_update_callback"))

    try:
        result = iyzico_client.initialize_card_update_checkout_form(
            subscription.external_customer_id, callback_url,
        )
    except iyzico_client.IyzicoError:
        messages.error(request, gettext("Couldn't open billing management. Please try again or contact us."))
        return redirect("billing_required")

    return render(request, "registration/billing_checkout.html", {
        "checkout_form_content": result["data"]["checkoutFormContent"],
        "heading": gettext("Update payment method"),
    })


@login_required
def billing_card_update_callback(request):
    messages.success(request, gettext("Your payment method has been updated."))
    return redirect("billing_required")


@csrf_exempt
def platform_webhook(request):
    if request.method != "POST":
        return HttpResponse(status=405)

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponse(status=400)

    event_type = payload.get("iyziEventType", "")
    subscription_ref = payload.get("subscriptionReferenceCode", "")
    order_ref = payload.get("orderReferenceCode", "")
    customer_ref = payload.get("customerReferenceCode", "")
    signature_header = request.headers.get("X-IYZ-SIGNATURE-V3", "")

    if not iyzico_client.verify_webhook_signature(
        settings.IYZICO_MERCHANT_ID, event_type, subscription_ref, order_ref, customer_ref, signature_header,
    ):
        return HttpResponse(status=401)

    subscription = Subscription.objects.filter(external_subscription_id=subscription_ref).first()
    if subscription is not None:
        if event_type == "subscription.order.success":
            subscription.status = Subscription.STATUS_ACTIVE
            subscription.save(update_fields=["status"])
        elif event_type == "subscription.order.failure":
            subscription.status = Subscription.STATUS_UNPAID
            subscription.save(update_fields=["status"])

    return HttpResponse(status=200)


# ---------------------------------------------------------------------------
# Customer views
# ---------------------------------------------------------------------------

@login_required
@permission_required("books.view_customer", raise_exception=True)
def customer_list(request):
    customers = Customer.objects.filter(account=request.account).annotate(
        invoice_count=Count("invoices")
    ).order_by("name")
    q = request.GET.get("q", "").strip()
    if q:
        customers = customers.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q))

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    paginator = Paginator(customers, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "books/customer_list.html", {
        "customers": page_obj.object_list,
        "page_obj": page_obj,
        "q": q,
        "query_string": query_string,
        "result_count_text": gettext("%(count)s customer(s) found") % {"count": paginator.count},
        "has_any_customers": Customer.objects.filter(account=request.account).exists(),
    })


@login_required
@permission_required("books.view_customer", raise_exception=True)
def customer_detail(request, id):
    customer = get_object_or_404(Customer, id=id, account=request.account)
    invoices = customer.invoices.filter(account=request.account).prefetch_related("items").order_by("-invoice_date")

    billed_by_currency = {}
    outstanding_by_currency = {}
    overdue_count = 0
    for invoice in invoices:
        billed_by_currency[invoice.currency] = billed_by_currency.get(invoice.currency, Decimal(0)) + invoice.grand_total
        if invoice.status != Invoice.STATUS_PAID:
            outstanding_by_currency[invoice.currency] = (
                outstanding_by_currency.get(invoice.currency, Decimal(0)) + invoice.grand_total
            )
        if invoice.is_overdue:
            overdue_count += 1

    paginator = Paginator(invoices, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    last_portal_login = (
        CustomerLoginToken.objects.filter(customer=customer, used_at__isnull=False)
        .aggregate(last=Max("used_at"))["last"]
    )

    return render(request, "books/customer_detail.html", {
        "customer": customer,
        "invoices": page_obj.object_list,
        "page_obj": page_obj,
        "billed_by_currency": sorted(billed_by_currency.items()),
        "outstanding_by_currency": sorted(outstanding_by_currency.items()),
        "overdue_count": overdue_count,
        "overdue_badge_text": gettext("%(count)s overdue") % {"count": overdue_count},
        "last_portal_login": last_portal_login,
    })


@login_required
@permission_required("books.add_customer", raise_exception=True)
def customer_create(request):
    form = CustomerForm()
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.owner = request.user
            customer.account = request.account
            customer.save()
            messages.success(request, gettext("Customer created."))
            return redirect("customer_list")
    return render(request, "books/customer_form.html", {"form": form})


@login_required
@permission_required("books.change_customer", raise_exception=True)
def customer_update(request, id):
    customer = get_object_or_404(Customer, id=id, account=request.account)
    form = CustomerForm(instance=customer)
    if request.method == "POST":
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, gettext("Customer updated."))
            return redirect("customer_list")
    return render(request, "books/customer_form.html", {"form": form, "customer": customer})


@login_required
@permission_required("books.delete_customer", raise_exception=True)
def customer_delete(request, id):
    customer = get_object_or_404(Customer, id=id, account=request.account)
    if request.method == "POST":
        customer.delete()
        messages.success(request, gettext("Customer deleted."))
        return redirect("customer_list")
    return render(request, "books/confirm_delete.html", {
        "object_type": gettext("customer"),
        "object_name": customer.name,
        "cancel_url": reverse("customer_list"),
    })


# ---------------------------------------------------------------------------
# Profit / Loss Report
# ---------------------------------------------------------------------------

def _pl_data(account, start_date, end_date):
    sales_qs = Sale.objects.filter(account=account)
    if start_date:
        sales_qs = sales_qs.filter(sale_date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(sale_date__lte=end_date)

    revenue_by_book = {
        item["book_id"]: item["revenue"] or 0
        for item in sales_qs.values("book_id").annotate(revenue=Sum(REVENUE_EXPRESSION))
    }
    units_by_book = {
        item["book_id"]: item["units"] or 0
        for item in sales_qs.values("book_id").annotate(units=Sum("quantity"))
    }

    returns_by_book = {
        item["sale__book_id"]: item["amount"] or 0
        for item in Return.objects.filter(sale__in=sales_qs)
        .values("sale__book_id")
        .annotate(amount=Sum(_RETURN_AMOUNT_EXPRESSION))
    }

    reorders_qs = Reorder.objects.filter(account=account, status=Reorder.STATUS_RECEIVED)
    if start_date:
        reorders_qs = reorders_qs.filter(received_at__date__gte=start_date)
    if end_date:
        reorders_qs = reorders_qs.filter(received_at__date__lte=end_date)
    purchase_by_book = {
        item["book_id"]: item["cost"] or 0
        for item in reorders_qs.values("book_id").annotate(cost=Sum(PURCHASE_COST_EXPRESSION))
    }

    book_ids = set(revenue_by_book)
    books = (
        Book.objects.filter(id__in=book_ids, account=account)
        .select_related("category")
        .order_by("title")
    )

    royalties_by_book = {}
    for rate in RoyaltyRate.objects.filter(book__in=books, account=account):
        rev = Decimal(str(revenue_by_book.get(rate.book_id, 0)))
        royalties_by_book[rate.book_id] = (
            royalties_by_book.get(rate.book_id, Decimal(0)) + rev * rate.rate / 100
        )

    rows = []
    totals = {k: Decimal(0) for k in ("revenue", "returns", "net_revenue", "purchase_cost", "distribution", "royalties", "net_profit")}

    for book in books:
        revenue = Decimal(str(revenue_by_book.get(book.id, 0)))
        returns = Decimal(str(returns_by_book.get(book.id, 0)))
        net_revenue = revenue - returns
        purchase_cost = Decimal(str(purchase_by_book.get(book.id, 0)))
        distribution = book.distribution_expense
        royalties = royalties_by_book.get(book.id, Decimal(0))
        net_profit = net_revenue - purchase_cost - distribution - royalties

        rows.append({
            "book": book,
            "units": units_by_book.get(book.id, 0),
            "revenue": revenue,
            "returns": returns,
            "net_revenue": net_revenue,
            "purchase_cost": purchase_cost,
            "distribution_expense": distribution,
            "royalties": royalties,
            "net_profit": net_profit,
        })

        totals["revenue"] += revenue
        totals["returns"] += returns
        totals["net_revenue"] += net_revenue
        totals["purchase_cost"] += purchase_cost
        totals["distribution"] += distribution
        totals["royalties"] += royalties
        totals["net_profit"] += net_profit

    return rows, totals


@login_required
@permission_required("books.view_book", raise_exception=True)
def profit_loss_report(request):
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    rows, totals = _pl_data(request.account, start_date, end_date)

    return render(request, "books/profit_loss_report.html", {
        "rows": rows,
        "start_date": start_date,
        "end_date": end_date,
        **totals,
    })


@login_required
@permission_required("books.view_book", raise_exception=True)
def export_profit_loss_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _register_pdf_fonts()
    body_font, bold_font = _pdf_fonts()

    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    rows, totals = _pl_data(request.account, start_date, end_date)

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = bold_font

    period = ""
    if start_date and end_date:
        period = f" ({start_date} – {end_date})"
    elif start_date:
        period = f" (from {start_date})"
    elif end_date:
        period = f" (to {end_date})"

    elements = [
        Paragraph(_pdf_text(gettext("Profit / Loss Report") + period), title_style),
        Spacer(1, 12),
    ]

    headers = [
        _pdf_text(gettext("Book")),
        _pdf_text(gettext("Units")),
        _pdf_text(gettext("Revenue")),
        _pdf_text(gettext("Returns")),
        _pdf_text(gettext("Net Revenue")),
        _pdf_text(gettext("Purchase Cost")),
        _pdf_text(gettext("Distribution")),
        _pdf_text(gettext("Royalties")),
        _pdf_text(gettext("Net Profit")),
    ]

    data = [headers]
    for row in rows:
        data.append([
            _pdf_text(row["book"].title),
            str(row["units"]),
            str(row["revenue"]),
            str(row["returns"]),
            str(row["net_revenue"]),
            str(row["purchase_cost"]),
            str(row["distribution_expense"]),
            str(row["royalties"]),
            str(row["net_profit"]),
        ])

    data.append([
        _pdf_text(gettext("TOTAL")),
        "",
        str(totals["revenue"]),
        str(totals["returns"]),
        str(totals["net_revenue"]),
        str(totals["purchase_cost"]),
        str(totals["distribution"]),
        str(totals["royalties"]),
        str(totals["net_profit"]),
    ])

    table = Table(data, colWidths=[140, 40, 65, 65, 70, 75, 70, 65, 70], repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f1f1f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), bold_font),
            ("FONTNAME", (0, 1), (-1, -2), body_font),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#1f1f1f")),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
            ("FONTNAME", (0, -1), (-1, -1), bold_font),
        ])
    )
    elements.append(table)

    document.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rumi-press-profit-loss.pdf"'
    return response


# ---------------------------------------------------------------------------
# Customer self-service portal
# ---------------------------------------------------------------------------

CUSTOMER_PORTAL_SESSION_KEY = "customer_portal_customer_id"


def _get_portal_customer(request):
    customer_id = request.session.get(CUSTOMER_PORTAL_SESSION_KEY)
    if not customer_id:
        return None
    return Customer.objects.filter(id=customer_id).first()


def customer_portal_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        customer = _get_portal_customer(request)
        if customer is None:
            return redirect("customer_portal_login")
        request.portal_customer = customer
        return view_func(request, *args, **kwargs)
    return wrapper


def _send_customer_portal_login_email(customer, link):
    send_mail(
        subject="Your RumiPress customer portal login link",
        message=(
            f"Hi {customer.name},\n\n"
            f"Click the link below to access your customer portal:\n{link}\n\n"
            f"This link expires in {settings.CUSTOMER_PORTAL_TOKEN_TTL_MINUTES} minutes "
            "and can only be used once.\n\n"
            "If you didn't request this, you can ignore this email."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[customer.email],
        fail_silently=True,
    )


def customer_portal_login(request):
    if _get_portal_customer(request):
        return redirect("customer_portal_dashboard")

    sent = False

    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        cache_key = f"customer_portal_login_cooldown:{email.lower()}"
        can_send = bool(email) and cache.add(cache_key, True, timeout=settings.CUSTOMER_PORTAL_LOGIN_COOLDOWN_SECONDS)

        for customer in Customer.objects.filter(email__iexact=email).exclude(email="") if can_send else []:
            login_token = CustomerLoginToken.objects.create(
                customer=customer,
                token=secrets.token_urlsafe(32),
                expires_at=timezone.now() + timedelta(minutes=settings.CUSTOMER_PORTAL_TOKEN_TTL_MINUTES),
            )
            link = request.build_absolute_uri(reverse("customer_portal_verify", args=[login_token.token]))
            _send_customer_portal_login_email(customer, link)
        sent = True

    return render(request, "customer_portal/login.html", {"sent": sent})


def customer_portal_verify(request, token):
    login_token = CustomerLoginToken.objects.select_related("customer").filter(token=token).first()

    if login_token is None or not login_token.is_valid:
        return render(request, "customer_portal/link_invalid.html", status=400)

    login_token.used_at = timezone.now()
    login_token.save(update_fields=["used_at"])

    request.session.flush()
    request.session[CUSTOMER_PORTAL_SESSION_KEY] = login_token.customer_id
    request.session.set_expiry(settings.CUSTOMER_PORTAL_SESSION_AGE_SECONDS)
    return redirect("customer_portal_dashboard")


@require_POST
def customer_portal_logout(request):
    request.session.flush()
    return redirect("customer_portal_login")


@customer_portal_required
def customer_portal_dashboard(request):
    customer = request.portal_customer
    invoices = customer.invoices.order_by("-invoice_date")

    billed_by_currency = {}
    outstanding_by_currency = {}
    overdue_count = 0
    for invoice in invoices:
        billed_by_currency[invoice.currency] = billed_by_currency.get(invoice.currency, Decimal(0)) + invoice.grand_total
        if invoice.status != Invoice.STATUS_PAID:
            outstanding_by_currency[invoice.currency] = (
                outstanding_by_currency.get(invoice.currency, Decimal(0)) + invoice.grand_total
            )
        if invoice.is_overdue:
            overdue_count += 1

    paginator = Paginator(invoices, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "customer_portal/dashboard.html", {
        "customer": customer,
        "invoices": page_obj.object_list,
        "page_obj": page_obj,
        "billed_by_currency": sorted(billed_by_currency.items()),
        "outstanding_by_currency": sorted(outstanding_by_currency.items()),
        "overdue_count": overdue_count,
        "overdue_badge_text": gettext("%(count)s overdue") % {"count": overdue_count},
    })


def _get_stripe_integration(account):
    return Integration.objects.filter(
        account=account, platform=Integration.PLATFORM_STRIPE, is_active=True,
    ).first()


@customer_portal_required
def customer_portal_invoice_detail(request, id):
    invoice = get_object_or_404(Invoice, id=id, customer=request.portal_customer)

    if request.GET.get("paid") == "1" and invoice.status != Invoice.STATUS_PAID:
        messages.success(
            request,
            gettext("Payment received — it may take a moment to reflect on this invoice."),
        )

    can_pay = (
        invoice.status != Invoice.STATUS_PAID
        and _get_stripe_integration(invoice.account) is not None
    )

    return render(request, "customer_portal/invoice_detail.html", {
        "invoice": invoice,
        "can_pay": can_pay,
    })


@customer_portal_required
def customer_portal_invoice_pdf(request, id):
    invoice = get_object_or_404(Invoice, id=id, customer=request.portal_customer)
    return _invoice_pdf_response(invoice)


@customer_portal_required
@require_POST
def customer_portal_invoice_pay(request, id):
    invoice = get_object_or_404(Invoice, id=id, customer=request.portal_customer)

    if invoice.status == Invoice.STATUS_PAID:
        return redirect("customer_portal_invoice_detail", id=invoice.id)

    integration = _get_stripe_integration(invoice.account)
    if integration is None:
        messages.error(request, gettext("Online payment isn't available for this invoice yet."))
        return redirect("customer_portal_invoice_detail", id=invoice.id)

    detail_url = request.build_absolute_uri(reverse("customer_portal_invoice_detail", args=[invoice.id]))
    client = stripe.StripeClient(api_key=integration.api_key)

    try:
        session = client.v1.checkout.sessions.create(params={
            "mode": "payment",
            "success_url": f"{detail_url}?paid=1",
            "cancel_url": detail_url,
            "metadata": {"invoice_id": str(invoice.id)},
            "line_items": [{
                "quantity": 1,
                "price_data": {
                    "currency": invoice.currency.lower(),
                    "unit_amount": int((invoice.grand_total * 100).to_integral_value()),
                    "product_data": {
                        "name": f"Invoice {invoice.invoice_number or invoice.id}",
                    },
                },
            }],
        })
    except stripe.StripeError:
        messages.error(request, gettext("Couldn't start payment. Please try again or contact us."))
        return redirect("customer_portal_invoice_detail", id=invoice.id)

    return redirect(session.url)
